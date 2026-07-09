from __future__ import annotations

import json
import gzip
import os
import pickle
import re
import shutil
import threading
import time
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from tkinter import BooleanVar, END, Label, StringVar, Tk, filedialog, messagebox, ttk

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageEnhance, ImageFilter, ImageGrab, ImageOps, ImageTk

try:
    import pytesseract
except Exception:  # pragma: no cover - handled at runtime for friendly UI
    pytesseract = None

try:
    from rapidfuzz import fuzz, process
except Exception:  # pragma: no cover - difflib fallback keeps the app usable
    fuzz = None
    process = None


def configure_ssl_certificates() -> None:
    import certifi

    default_bundle = certifi.where()
    for var in ("SSL_CERT_FILE", "REQUESTS_CA_BUNDLE", "CURL_CA_BUNDLE"):
        value = os.environ.get(var)
        if value and not Path(value).is_file():
            os.environ[var] = default_bundle


configure_ssl_certificates()


APP_DIR = Path(__file__).resolve().parent
APP_VERSION = "2026-07-08.17"
DATA_DIR = APP_DIR / "data"
SCRYFALL_CARDS_PATH = DATA_DIR / "scryfall_default_cards.json"
SETS_PATH = DATA_DIR / "scryfall_sets.json"
MTGJSON_ATOMIC_PATH = DATA_DIR / "mtgjson_atomic_cards.json.gz"
INDEX_PATH = DATA_DIR / "card_index.pkl"
INDEX_VERSION = 24
# Latin-script languages whose printed names (from foreign-only printings in
# the Scryfall bulk, e.g. FBB French "Lunettes d'Urza") join the name pools.
_FOREIGN_NAME_LANGS = {"fr", "es", "it", "de", "pt"}
# Layouts that reuse real card names but are never scanned as collectible cards;
# excluded from the fuzzy name/face matching pools.
_NON_CATALOG_LAYOUTS = {"art_series", "token", "double_faced_token", "emblem", "reversible_card"}
EXCEL_PATH = APP_DIR / "MagicCollection.xlsx"
DEBUG_PATH = APP_DIR / "last_extraction_debug.txt"
DEBUG_LOG_PATH = APP_DIR / "extraction_log.txt"
HTTP_HEADERS = {
    "User-Agent": "magic-extractor/1.0 (local desktop tool)",
    "Accept": "application/json",
}

FIELDS = [
    "Nome",
    "Tipo",
    "Raridade",
    "Descrição magia",
    "Custo mana",
    "Cor(cores)",
    "Preço mínimo",
    "Coleção",
    "É foil?",
    "Ano",
    "Poder",
    "Qtd",
]

RARITY_PT = {
    "common": "Comum",
    "uncommon": "Incomum",
    "rare": "Rara",
    "mythic": "Mítica",
    "special": "Especial",
    "bonus": "Bônus",
}

RARITY_EN = {
    "common": "Common",
    "uncommon": "Uncommon",
    "rare": "Rare",
    "mythic": "Mythic",
    "special": "Special",
    "bonus": "Bonus",
}

COLOR_PT = {
    "W": "Branco",
    "U": "Azul",
    "B": "Preto",
    "R": "Vermelho",
    "G": "Verde",
}

TESSERACT_CANDIDATES = [
    r"C:\Program Files\Tesseract-OCR\tesseract.exe",
    r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
    str(Path(os.environ.get("LOCALAPPDATA", "")) / "Programs" / "Tesseract-OCR" / "tesseract.exe"),
]

MANUAL_PT_TRANSLATIONS = {
    "Dwarven Warriors": {
        "printed_name": "An\u00f5es Guerreiros",
        "printed_type_line": "Criatura \u2014 An\u00e3o",
        "printed_text": (
            "{T}: A criatura-alvo com poder igual ou inferior a 2 n\u00e3o pode ser bloqueada at\u00e9 o final do turno. "
            "Outros efeitos podem ser usados depois para aumentar o poder da criatura acima de 2."
        ),
        "preferred_set": "4ED",
    },
    "Joven": {
        "printed_name": "Joven",
        "printed_type_line": "Criatura Lend\u00e1ria \u2014 Humano Ladino",
        "printed_text": "{R}{R}{R}, {T}: Destr\u00f3i o artefato alvo que n\u00e3o seja criatura.",
        "preferred_set": "HML",
    },
    "Earth Elemental": {
        "printed_name": "Elemental da Terra",
        "printed_type_line": "Criatura \u2014 Elemental",
        "printed_text": "",
        "preferred_set": "4ED",
    },
    "Maze Glider": {
        "printed_name": "Planador do Labirinto",
        "printed_type_line": "Criatura — Elemental",
        "printed_text": "Voar\nAs criaturas multicoloridas que você controla têm voar.",
    },
    "Opal Lake Gatekeepers": {
        "printed_name": "Porteiros do Lago de Opala",
        "printed_type_line": "Criatura — Vedalkeano Soldado",
        "printed_text": (
            "Quando Porteiros do Lago de Opala entrar no campo de batalha, "
            "se você controlar dois ou mais Portões, poderá comprar um card."
        ),
    },
    "Murmuring Phantasm": {
        "printed_name": "Fantasma Murmurante",
        "printed_type_line": "Criatura — Espírito",
        "printed_text": "Defensor",
    },
    "Immolation": {
        "printed_name": "Imola\u00e7\u00e3o",
        "printed_type_line": "Encantar Criatura",
        "printed_text": "A criatura-alvo ganha +2/-2.",
        "preferred_set": "4ED",
    },
    "The Brute": {
        "printed_name": "O Bruto",
        "printed_type_line": "Encantar Criatura",
        "printed_text": (
            "A criatura-alvo ganha +1/+0.\n"
            "{R}{R}{R}: Regenera a criatura-alvo encantada pelo Bruto."
        ),
        "preferred_set": "4ED",
    },
    "Flare": {
        "printed_name": "Detona\u00e7\u00e3o",
        "printed_type_line": "M\u00e1gica Instant\u00e2nea",
        "printed_text": (
            "Detona\u00e7\u00e3o causa 1 ponto de dano a qualquer alvo. "
            "Compre um card no in\u00edcio da manuten\u00e7\u00e3o do pr\u00f3ximo turno."
        ),
    },
    "Shatter": {
        "printed_name": "Estilha\u00e7ar",
        "printed_type_line": "M\u00e1gica Instant\u00e2nea",
        "printed_text": "Destr\u00f3i artefato-alvo.",
        "preferred_set": "4ED",
    },
    "Tunnel": {
        "printed_name": "T\u00fanel",
        "printed_type_line": "M\u00e1gica Instant\u00e2nea",
        "printed_text": "Enterra a barreira-alvo.",
        "preferred_set": "4ED",
    },
    "Warrant // Warden": {
        "printed_name": "Permiss\u00e3o // Prote\u00e7\u00e3o",
        "printed_type_line": "M\u00e1gica Instant\u00e2nea // Feiti\u00e7o",
        "printed_text": (
            "Coloque a criatura alvo atacante ou bloqueadora no topo do grim\u00f3rio de seu dono. // "
            "Crie uma ficha de criatura branca e azul 4/4 do tipo Esfinge com voar e vigil\u00e2ncia."
        ),
    },
    "Flotsam // Jetsam": {
        "printed_name": "Destro\u00e7os // Refugo",
        "printed_type_line": "M\u00e1gica Instant\u00e2nea // Feiti\u00e7o",
        "printed_text": (
            "Triture tr\u00eas cards. Investigue. (Crie uma ficha de Pista. Ela \u00e9 um artefato com "
            "\"{2}, sacrifique este artefato: Compre um card.\") // "
            "Cada oponente tritura tr\u00eas cards, depois voc\u00ea pode conjurar uma m\u00e1gica do cemit\u00e9rio "
            "de cada oponente sem pagar seu custo de mana. Se uma m\u00e1gica conjurada dessa forma seria "
            "colocada em um cemit\u00e9rio, em vez disso exile-a."
        ),
    },
    "Volshe Tideturner": {
        "printed_name": "Inversora de Mar\u00e9s Volshe",
        "printed_type_line": "Criatura \u2014 Trit\u00e3o Mago",
        "printed_text": (
            "{T}: Adicione {U}. Gaste este mana somente para conjurar uma m\u00e1gica instant\u00e2nea, "
            "um feiti\u00e7o ou uma m\u00e1gica refor\u00e7ada."
        ),
    },
}

PREFERRED_PRINT_SETS = {
    "Dwarven Warriors": "4ED",
    "Joven": "HML",
    "Earth Elemental": "4ED",
    "Dwarven Blastminer": "ONS",
    "Spikeshot Goblin": "MRD",
}


def normalize_text(value: str) -> str:
    value = unicodedata.normalize("NFKD", value or "")
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = value.lower()
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def clean_ocr_line(line: str) -> str:
    line = line.replace("|", "I").strip()
    line = re.sub(r"\s+", " ", line)
    line = re.sub(r"[\{\}\[\]]", "", line)
    return line.strip(" .,:;")


COMMON_PT_OCR_WORDS = {
    "cemiterio",
    "cemiterios",
    "grimorio",
    "grimorios",
    "conjura",
    "conjurar",
    "magica",
    "magicas",
    "magia",
    "magias",
    "criatura",
    "criaturas",
    "terreno",
    "terrenos",
    "oponente",
    "oponentes",
    "jogador",
    "jogadores",
    "card",
    "cards",
    "turno",
    "turnos",
    "poder",
    "resistencia",
    "controle",
    "alvo",
    "feiticaria",
    "encantamento",
    "instantanea",
    "instantaneas",
    "revela",
    "revelar",
    "coloque",
    "comprar",
    "compre",
    "exila",
    "exile",
    "dano",
    "combate",
    "etapa",
    "final",
    "inicio",
    "proprio",
    "propria",
    "numero",
    "cada",
    "vezes",
    "vez",
    "quando",
    "voce",
    "seus",
    "suas",
    "seu",
    "sua",
    "ate",
    "todo",
    "toda",
    "todos",
    "todas",
    "com",
    "para",
    "mais",
    "menos",
    "outro",
    "outra",
    "outros",
    "outras",
}

COMMON_EN_OCR_WORDS = {
    "card",
    "cards",
    "draw",
    "land",
    "lands",
    "damage",
    "creature",
    "creatures",
    "target",
    "opponent",
    "opponents",
    "control",
    "enters",
    "whenever",
    "ability",
    "turn",
    "spell",
    "spells",
    "player",
    "players",
    "graveyard",
    "library",
    "battlefield",
    "permanent",
    "permanents",
    "counter",
    "counters",
    "token",
    "tokens",
    "serving",
    "nothing",
    "like",
    "posters",
    "made",
    "second",
    "resolved",
    "legendary",
    "pilot",
}

MTG_KEYWORD_NAME_STOPWORDS = {
    "deathtouch",
    "defender",
    "first strike",
    "flash",
    "flying",
    "haste",
    "hexproof",
    "indestructible",
    "lifelink",
    "menace",
    "reach",
    "trample",
    "vigilance",
    "alcance",
    "ameaca",
    "atropelar",
    "defensor",
    "impeto",
    "indestrutivel",
    "iniciativa",
    "lampejo",
    "toque mortifero",
    "vigilancia",
    "voar",
}

AMBIGUOUS_SET_CODE_WORDS = {
    "A",
    "AND",
    "ARE",
    "AS",
    "COM",
    "DAS",
    "DE",
    "DOS",
    "E",
    "EM",
    "MED",
    "NEM",
    "NO",
    "POR",
    "STA",
    "THE",
    "UMA",
    "VOC",
    # Real set codes that are also common English words in rules text; they
    # only count as print evidence in true footer shape (language code next).
    "ALL",
    "BOT",
    "CON",
    "DIS",
    "DOM",
    "EVE",
    "ICE",
    "MAT",
    "ONE",
    "WAR",
    "WHO",
}


def is_plausible_namebar(text: str) -> bool:
    cleaned = clean_ocr_line(text or "")
    words = re.findall(r"[^\W\d_]{3,}", cleaned, flags=re.UNICODE)
    if not words or len(words) > 6:
        return False
    if not any(len(word) >= 4 for word in words):
        return False
    alpha = sum(ch.isalpha() for ch in cleaned)
    if cleaned and alpha < len(cleaned) * 0.35:
        return False
    title_words = re.findall(r"\b[^\W\d_]{3,}\b", cleaned, flags=re.UNICODE)
    if not title_words and len(words) <= 3:
        return False
    return True


def _namebar_line_score(line: str) -> int:
    score = sum(1 for word in line.split() if len(word) >= 4)
    if "," in line:
        score += 5
    if re.match(r"^[^\W\d_]", line, flags=re.UNICODE):
        score += 2
    return score


def ranked_namebar_lines(text: str) -> list[str]:
    lines = [clean_ocr_line(line) for line in (text or "").splitlines()]
    candidates = [line for line in lines if is_plausible_namebar(line)]
    ranked = sorted(candidates, key=_namebar_line_score, reverse=True)
    seen = set()
    unique = []
    for line in ranked:
        key = normalize_text(line)
        if key and key not in seen:
            seen.add(key)
            unique.append(line)
    return unique


def best_namebar_line(text: str) -> str:
    ranked = ranked_namebar_lines(text)
    return ranked[0] if ranked else ""


def normalize_ocr_name(value: str) -> str:
    value = clean_ocr_line(value or "")
    tokens = []
    for token in value.split():
        if re.search(r"\d", token) and re.search(r"[A-Za-z]", token):
            token = token.translate(str.maketrans({"1": "l", "0": "o", "5": "s", "8": "b"}))
        tokens.append(token)
    return " ".join(tokens)


def clean_card_text(value: str) -> str:
    value = value or ""
    value = value.replace("\r\n", "\n").replace("\r", "\n")
    value = re.sub(r"\s*\n\s*", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def title_words(value: str) -> list[str]:
    return [word for word in normalize_text(value).split() if len(word) >= 3 or word in {"a", "an", "of", "to"}]


def collector_key(value: str) -> str:
    value = str(value or "").upper().strip()
    match = re.match(r"0*(\d+)([A-Z]*)$", value)
    if match:
        return f"{int(match.group(1))}{match.group(2)}"
    return value


def ocr_collector_key(value: str) -> str:
    value = str(value or "").upper().strip()
    value = value.translate(str.maketrans({"O": "0", "I": "1", "L": "1"}))
    value = re.sub(r"^[568](\d{3}[A-Z]?)$", r"\1", value)
    return collector_key(value)


def normalized_phrase_in_text(normalized_name: str, normalized_raw: str) -> bool:
    """Word-boundary containment: "lich" must not match inside "willich"."""
    if not normalized_name:
        return False
    return bool(re.search(rf"\b{re.escape(normalized_name)}\b", normalized_raw))


def words_fuzzy_match(expected: str, word: str, threshold: int = 82) -> bool:
    if expected == word:
        return True
    if not fuzz:
        return False
    length_gap = abs(len(expected) - len(word))
    effective_threshold = threshold
    if length_gap >= 3:
        effective_threshold = max(threshold, 90)
    elif length_gap >= 2:
        effective_threshold = max(threshold, 88)
    ratio = fuzz.ratio(expected, word)
    if ratio < effective_threshold:
        return False
    if length_gap >= 3 and min(len(expected), len(word)) >= 5:
        shorter = expected if len(expected) < len(word) else word
        longer = word if shorter == expected else expected
        if longer.startswith(shorter[: max(4, len(shorter))]):
            return ratio >= 92
    return True


def fuzzy_word_count(expected_words: list[str], text_words: list[str], threshold: int = 82) -> int:
    count = 0
    for expected in expected_words:
        if any(words_fuzzy_match(expected, word, threshold) for word in text_words):
            count += 1
    return count


def detect_tesseract_languages() -> tuple[str, bool]:
    if pytesseract is None:
        return "eng", False
    try:
        langs = pytesseract.get_languages(config="")
        if "por" in langs:
            return "por+eng", True
    except Exception:
        pass
    return "eng", False


def bilingual(en_value: str, pt_value: str) -> str:
    en_value = clean_card_text(en_value)
    pt_value = clean_card_text(pt_value)
    if pt_value and normalize_text(pt_value) != normalize_text(en_value):
        return f"{en_value} / {pt_value}" if en_value else pt_value
    return en_value


def split_face_values(value: str) -> list[str]:
    return [part.strip() for part in re.split(r"\s*//\s*", value or "") if part.strip()]


def bilingual_name_parts(value: str) -> set[str]:
    value = clean_card_text(value)
    parts = [part.strip() for part in value.split(" / ") if part.strip()]
    expanded_parts = []
    for part in parts:
        expanded_parts.append(part)
        expanded_parts.extend(split_face_values(part))
    normalized = {normalize_text(part) for part in expanded_parts}
    if len(parts) > 1:
        normalized.add(normalize_text(value))
    return {part for part in normalized if part}


def excel_names_match(cell_value: str, search_name: str) -> bool:
    cell_parts = bilingual_name_parts(cell_value)
    search_parts = bilingual_name_parts(search_name)
    if not cell_parts or not search_parts:
        return False
    return bool(cell_parts & search_parts)


def find_name_row_in_excel(name: str) -> int | None:
    if not EXCEL_PATH.exists():
        return None
    workbook = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        for row_index, row in enumerate(
            sheet.iter_rows(min_col=1, max_col=1, values_only=True),
            start=1,
        ):
            cell_value = row[0] if row else None
            if row_index == 1 and cell_value and str(cell_value).strip() == "Nome":
                continue
            if cell_value is None or str(cell_value).strip() == "":
                continue
            if excel_names_match(str(cell_value), name):
                return row_index
    finally:
        workbook.close()
    return None


def ensure_excel_workbook():
    if EXCEL_PATH.exists():
        workbook = load_workbook(EXCEL_PATH)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Cartas"
        sheet.append(FIELDS)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
        for index, field in enumerate(FIELDS, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = max(12, min(42, len(field) + 8))
    return workbook, sheet


def append_row_to_excel(row: list[str]) -> None:
    workbook, sheet = ensure_excel_workbook()
    sheet.append(row)
    workbook.save(EXCEL_PATH)


def update_row_in_excel(row_number: int, row: list[str]) -> None:
    workbook = load_workbook(EXCEL_PATH)
    sheet = workbook.active
    for col_index, value in enumerate(row, start=1):
        sheet.cell(row=row_number, column=col_index, value=value)
    workbook.save(EXCEL_PATH)


def card_faces_text(card: dict, field: str) -> str:
    if card.get(field):
        return card[field]
    faces = card.get("card_faces") or []
    parts = []
    for face in faces:
        value = face.get(field)
        if value:
            name = face.get("name", "")
            parts.append(f"{name}: {value}" if name else value)
    return " // ".join(parts)


def card_power(card: dict) -> str:
    if card.get("power") and card.get("toughness"):
        return f"{card['power']}/{card['toughness']}"
    for face in card.get("card_faces") or []:
        if face.get("power") and face.get("toughness"):
            return f"{face['power']}/{face['toughness']}"
    return ""


def card_mana_cost(card: dict) -> str:
    if card.get("mana_cost"):
        return card["mana_cost"]
    costs = [face.get("mana_cost", "") for face in card.get("card_faces") or []]
    return " // ".join(cost for cost in costs if cost)


def card_colors_pt(card: dict) -> str:
    colors = card.get("colors") or card.get("color_identity") or []
    if not colors:
        return "Incolor"
    return ", ".join(COLOR_PT.get(color, color) for color in colors)


def crop_possible_app_screenshot(image: Image.Image) -> Image.Image:
    width, height = image.size
    if width / max(1, height) <= 1.25:
        return image
    preview_width = min(int(width * 0.45), int(height * 0.62))
    top = int(height * 0.38)
    crop = image.crop((0, top, preview_width, height))
    card_crop = crop_card_preview(crop)
    if card_crop is not None:
        return card_crop
    if crop.width >= 160 and crop.height >= 240:
        return crop
    return image


def crop_card_preview(image: Image.Image) -> Image.Image | None:
    image = image.convert("RGB")
    width, height = image.size
    dark_rows = []
    dark_cols = []
    row_threshold = max(12, int(width * 0.08))
    col_threshold = max(12, int(height * 0.08))

    for y in range(height):
        count = 0
        for x in range(width):
            r, g, b = image.getpixel((x, y))
            if r + g + b < 120:
                count += 1
        if count >= row_threshold:
            dark_rows.append(y)

    for x in range(width):
        count = 0
        for y in range(height):
            r, g, b = image.getpixel((x, y))
            if r + g + b < 120:
                count += 1
        if count >= col_threshold:
            dark_cols.append(x)

    if not dark_rows or not dark_cols:
        return None

    left = max(0, min(dark_cols) - 10)
    right = min(width, max(dark_cols) + 11)
    top = max(0, min(dark_rows) - 10)
    bottom = min(height, max(dark_rows) + 40)
    crop_width = right - left
    crop_height = bottom - top
    if crop_width < 120 or crop_height < 180:
        return None
    ratio = crop_width / max(1, crop_height)
    if not 0.45 <= ratio <= 0.9:
        return None
    return image.crop((left, top, right, bottom))


def configure_tesseract() -> str | None:
    if pytesseract is None:
        return None
    found = shutil.which("tesseract")
    if found:
        pytesseract.pytesseract.tesseract_cmd = found
        return found
    for candidate in TESSERACT_CANDIDATES:
        if Path(candidate).exists():
            pytesseract.pytesseract.tesseract_cmd = candidate
            return candidate
    return None


@dataclass
class OcrMatchContext:
    ocr: OcrResult
    fractions: list[tuple[str, int]]
    tokens: list[str]
    sets_in_text: set[str]
    hint_card: dict | None = None
    namebar_resolved: bool = False
    hint_from_title: bool = False


@dataclass
class OcrResult:
    name_hint: str = ""
    namebar_text: str = ""
    set_code: str = ""
    collector_number: str = ""
    rarity_hint: str = ""
    raw_text: str = ""


class LocalOcr:
    def __init__(self) -> None:
        self.tesseract_path = configure_tesseract()
        self.ocr_lang, self.por_available = detect_tesseract_languages()

    def available(self) -> bool:
        return pytesseract is not None and self.tesseract_path is not None

    def extract(self, image: Image.Image) -> OcrResult:
        if not self.available():
            raise RuntimeError(
                "Tesseract OCR não foi encontrado. Instale o Tesseract e reinicie o app."
            )

        image = crop_possible_app_screenshot(image.convert("RGB"))
        texts = []
        namebar_texts = []
        for candidate in self._candidate_images(image):
            width, height = candidate.size
            name_crop = candidate.crop((0, int(height * 0.035), int(width * 0.84), int(height * 0.14)))
            top_name_crop = candidate.crop(
                (
                    int(width * 0.02),
                    int(height * 0.015),
                    int(width * 0.75),
                    int(height * 0.095),
                )
            )
            focused_name_crop = candidate.crop(
                (
                    int(width * 0.055),
                    int(height * 0.055),
                    int(width * 0.48),
                    int(height * 0.12),
                )
            )
            bottom_crop = candidate.crop((0, int(height * 0.86), width, height))
            footer_crop = candidate.crop((0, int(height * 0.74), width, height))
            namebar_texts.extend(
                [
                    self._ocr(name_crop, "--psm 7", bilingual=True),
                    self._ocr(top_name_crop, "--psm 7", bilingual=True),
                    self._ocr(focused_name_crop, "--psm 7", bilingual=True),
                    self._ocr_dark_namebar(top_name_crop),
                    self._ocr(candidate.crop((0, 0, int(width * 0.9), int(height * 0.18))), "--psm 6", bilingual=True),
                ]
            )
            texts.extend(
                [
                    self._ocr(name_crop, "--psm 7", bilingual=True),
                    self._ocr(top_name_crop, "--psm 7", bilingual=True),
                    self._ocr(focused_name_crop, "--psm 7", bilingual=True),
                    self._ocr_dark_namebar(top_name_crop),
                    self._ocr(bottom_crop, "--psm 6", bilingual=True),
                    self._ocr_footer(footer_crop),
                    self._ocr(candidate, "--psm 6"),
                ]
            )

        raw_text = "\n".join(part for part in texts if part.strip())
        namebar_text = "\n".join(part for part in namebar_texts if part.strip())
        name_hint = self._parse_name(raw_text)
        set_code, collector_number, rarity_hint = self._parse_collector(raw_text)
        return OcrResult(
            name_hint=name_hint,
            namebar_text=namebar_text,
            set_code=set_code,
            collector_number=collector_number,
            rarity_hint=rarity_hint,
            raw_text=raw_text,
        )

    @staticmethod
    def _candidate_images(image: Image.Image) -> list[Image.Image]:
        width, height = image.size
        if width / max(1, height) > 0.9:
            candidates = []
            preview_width = min(int(width * 0.45), int(height * 0.62))
            boxes = [
                (0, int(height * 0.38), preview_width, height),
                (0, int(height * 0.45), preview_width, height),
                (0, 0, preview_width, height),
            ]
            for box in boxes:
                crop = image.crop(box)
                if crop.width > 120 and crop.height > 180:
                    candidates.append(crop)
            return candidates
        candidates = [image]
        return candidates

    def _ocr(self, image: Image.Image, config: str, bilingual: bool = False) -> str:
        prepared = self._prepare(image)
        lang = self.ocr_lang if bilingual and self.por_available else "eng"
        return pytesseract.image_to_string(prepared, lang=lang, config=config)

    def _ocr_dark_namebar(self, image: Image.Image) -> str:
        grayscale = image.convert("L")
        scale = max(6, int(1800 / max(1, grayscale.width)))
        grayscale = grayscale.resize((grayscale.width * scale, grayscale.height * scale))
        lang = self.ocr_lang if self.por_available else "eng"
        results = []
        for threshold in (180, 200):
            binary = grayscale.point(lambda pixel, limit=threshold: 255 if pixel > limit else 0)
            results.append(pytesseract.image_to_string(binary, lang=lang, config="--psm 7"))
        return "\n".join(results)

    def _ocr_footer(self, image: Image.Image) -> str:
        original_footer = image.convert("L")
        footer = original_footer
        footer = footer.resize((footer.width * 5, footer.height * 5))
        simple_binary = footer.point(lambda pixel: 255 if pixel > 100 else 0)
        contrast_footer = ImageEnhance.Contrast(footer).enhance(2.5)
        binary = contrast_footer.point(lambda pixel: 255 if pixel > 100 else 0)
        inverted_binary = ImageOps.invert(contrast_footer).point(lambda pixel: 255 if pixel > 100 else 0)
        lower_footer = original_footer.crop((0, int(original_footer.height * 0.54), original_footer.width, original_footer.height))
        lower_footer = lower_footer.resize((lower_footer.width * 6, lower_footer.height * 6))
        lower_contrast = ImageEnhance.Contrast(lower_footer).enhance(2.5)
        lower_inverted = ImageOps.invert(lower_contrast).point(lambda pixel: 255 if pixel > 110 else 0)
        copyright_footer = original_footer.crop(
            (0, int(original_footer.height * 0.56), int(original_footer.width * 0.95), original_footer.height)
        )
        copyright_footer = copyright_footer.resize((copyright_footer.width * 8, copyright_footer.height * 8))
        copyright_contrast = ImageEnhance.Contrast(copyright_footer).enhance(3.0)
        copyright_sharp = ImageEnhance.Contrast(copyright_footer.filter(ImageFilter.SHARPEN)).enhance(3.0)
        copyright_light = copyright_contrast.point(lambda pixel: 255 if pixel > 70 else 0)
        copyright_binary = copyright_contrast.point(lambda pixel: 255 if pixel > 80 else 0)
        copyright_inverted = ImageOps.invert(copyright_contrast).point(lambda pixel: 255 if pixel > 170 else 0)
        copyright_sharp_binary = copyright_sharp.point(lambda pixel: 255 if pixel > 110 else 0)
        wide_copyright = original_footer.crop(
            (0, int(original_footer.height * 0.52), int(original_footer.width * 0.98), original_footer.height)
        )
        wide_copyright = wide_copyright.resize((wide_copyright.width * 8, wide_copyright.height * 8))
        wide_copyright = ImageEnhance.Contrast(wide_copyright).enhance(3.0)
        wide_copyright_light = wide_copyright.point(lambda pixel: 255 if pixel > 70 else 0)
        low_copyright = original_footer.crop(
            (0, int(original_footer.height * 0.64), int(original_footer.width * 0.98), original_footer.height)
        )
        low_copyright = low_copyright.resize((low_copyright.width * 8, low_copyright.height * 8))
        low_copyright = ImageEnhance.Contrast(low_copyright).enhance(3.0)
        low_copyright_light = low_copyright.point(lambda pixel: 255 if pixel > 70 else 0)
        low_copyright_binary = low_copyright.point(lambda pixel: 255 if pixel > 90 else 0)
        lang = self.ocr_lang if self.por_available else "eng"
        return "\n".join(
            [
                pytesseract.image_to_string(simple_binary, lang=lang, config="--psm 6"),
                pytesseract.image_to_string(binary, lang=lang, config="--psm 6"),
                pytesseract.image_to_string(inverted_binary, lang=lang, config="--psm 6"),
                pytesseract.image_to_string(lower_inverted, lang=lang, config="--psm 6"),
                pytesseract.image_to_string(copyright_light, lang=lang, config="--psm 6"),
                pytesseract.image_to_string(copyright_binary, lang=lang, config="--psm 6"),
                pytesseract.image_to_string(copyright_inverted, lang=lang, config="--psm 6"),
                pytesseract.image_to_string(copyright_sharp_binary, lang=lang, config="--psm 6"),
                pytesseract.image_to_string(wide_copyright_light, lang=lang, config="--psm 6"),
                pytesseract.image_to_string(low_copyright_light, lang=lang, config="--psm 6"),
                pytesseract.image_to_string(low_copyright_binary, lang=lang, config="--psm 6"),
            ]
        )

    @staticmethod
    def _prepare(image: Image.Image) -> Image.Image:
        image = image.convert("L")
        scale = max(2, int(1200 / max(1, image.width)))
        image = image.resize((image.width * scale, image.height * scale))
        image = ImageEnhance.Contrast(image).enhance(1.8)
        return image.point(lambda pixel: 255 if pixel > 155 else 0)

    @staticmethod
    def _parse_name(text: str) -> str:
        lines = [clean_ocr_line(line) for line in text.splitlines()]
        lines = [line for line in lines if len(line) >= 3]
        if not lines:
            return ""
        ignored = {
            "colar",
            "imagem",
            "abrir",
            "extrair",
            "foil",
            "copiar",
            "google",
            "docs",
            "cabecalho",
            "salvar",
            "excel",
            "atualizar",
            "base",
            "nome",
            "tipo",
            "raridade",
            "descricao",
            "magia",
            "custo",
            "mana",
            "cor",
            "preco",
            "minimo",
            "colecao",
            "ano",
            "poder",
            "qtd",
            "ocr",
            "erro",
            "changeling",
            "this",
            "card",
            "every",
            "creature",
            "type",
            "is",
        }

        def score(candidate: str) -> int:
            normalized = normalize_text(candidate)
            words = normalized.split()
            if not words or all(word in ignored for word in words):
                return -100
            title_words = re.findall(r"\b[^\W\d_]{3,}\b", candidate, flags=re.UNICODE)
            value = min(len(words), 5)
            value += len(title_words) * 3
            if len(title_words) >= 2:
                value += 10
            elif not title_words:
                value -= 6
            value -= sum(3 for word in words if word in ignored)
            value -= sum(2 for word in words if word in {"instant", "sorcery", "artifact", "enchantment", "land"})
            value -= sum(1 for word in words if word.isdigit())
            if re.search(
                r"[©™]|\b(?:h?illus|iilus|iilug|inus|tilus|tilug|tl?lus|tlug|tlust|titus|hust|mlus[et]?|mus)\b|"
                r"wizards|coast|wasatch|anthony|palumbo|paliso",
                candidate,
                re.IGNORECASE,
            ):
                value -= 50
            if "/" in candidate or re.search(r"\b\d{3,}\b", candidate):
                value -= 25
            if len(words) > 6:
                value -= 15
            if re.search(r"\b(criatura|creature|instantanea|feiticaria|encantamento|terreno|artifact)\b", normalized):
                value -= 30
            if re.search(r"\b(da|de|do|das|dos)\b", normalized) and title_words:
                value += 8
            if re.search(r"\b(was|were|nothing|serving|like the|made it|whenever a)\b", normalized):
                value -= 45
            if re.search(r"\b(the|and|for|with|that|this|from|into)\b", normalized) and len(words) >= 5:
                value -= 20
            if "," in candidate:
                value += 12
            if re.match(r"^[^\W\d_]+,", candidate, flags=re.UNICODE):
                value += 15
            if re.search(r"[^\w\s\-',]", candidate, flags=re.UNICODE):
                value -= 10
            return value

        line = max(lines, key=score)
        line = re.sub(r"\s+[\dXWUBRGC/]+$", "", line, flags=re.IGNORECASE)
        return clean_ocr_line(line)

    @staticmethod
    def _parse_collector(text: str) -> tuple[str, str, str]:
        text = text.upper().replace("•", " ")
        text = re.sub(r"\bO(\d{2,4})\b", r"0\1", text)
        text = re.sub(r"[^A-Z0-9]+", " ", text)
        long_footer_candidates = []
        for match in re.finditer(
            r"\b([CMUR])(?:\1)?\s+(0*\d{3,4}[A-Z]?)\b[\s\S]{0,40}?\b([A-Z]{3})\b[\s\S]{0,20}?\b(?:EN|PT)\b",
            text,
        ):
            rarity, number, set_code = match.groups()
            normalized_number = ocr_collector_key(number)
            if len(re.sub(r"\D", "", normalized_number)) >= 3:
                long_footer_candidates.append((set_code, normalized_number, rarity, match.start()))
        if long_footer_candidates:
            set_code, number, rarity, _position = max(
                long_footer_candidates,
                key=lambda item: (len(re.sub(r"\D", "", item[1])), -item[3]),
            )
            return set_code, number, rarity
        rarity = ""
        rarity_with_pt_match = re.search(
            r"\b([CMUR])\s+(\d{1,4}[A-Z]?)\s+(?:\d+\s+\d+\s+)?([A-Z0-9]{2,5})\s+(?:EN|PT|ES|FR|DE|IT|JP|KO|RU|ZHS|ZHT)\b",
            text,
        )
        if rarity_with_pt_match:
            rarity, number, set_code = rarity_with_pt_match.groups()
            return set_code, ocr_collector_key(number), rarity

        rarity_match = re.search(
            r"\b([CMUR])\s+(\d{1,4}[A-Z]?)\s+([A-Z0-9]{2,5})\s+(?:EN|PT|ES|FR|DE|IT|JP|KO|RU|ZHS|ZHT)\b",
            text,
        )
        if rarity_match:
            rarity, number, set_code = rarity_match.groups()
            return set_code, ocr_collector_key(number), rarity

        loose_footer = re.search(
            r"\b([CMUR])\s+(0*\d{1,4}[A-Z]?)\b[\s\S]{0,40}?\b([A-Z]{3})\b[\s\S]{0,20}?\b(?:EN|PT)\b",
            text,
        )
        if loose_footer:
            rarity, number, set_code = loose_footer.groups()
            return set_code, ocr_collector_key(number), rarity

        match = re.search(r"\b(\d{1,4}[A-Z]?)\s+([A-Z0-9]{2,5})\s+(?:EN|PT|ES|FR|DE|IT|JP|KO|RU|ZHS|ZHT)\b", text)
        if match:
            number, set_code = match.groups()
            return set_code, ocr_collector_key(number), rarity

        return "", "", ""


class ScryfallDatabase:
    def __init__(self) -> None:
        self.cards: list[dict] = []
        self.sets: dict[str, dict] = {}
        self.by_print: dict[tuple[str, str], list[dict]] = {}
        self.pt_by_oracle: dict[str, list[dict]] = {}
        self.en_by_oracle: dict[str, list[dict]] = {}
        self.mtgjson_pt_by_name: dict[str, dict] = {}
        self.name_choices: list[tuple[str, int]] = []
        self.pt_name_choices: list[tuple[str, int]] = []
        self.face_name_choices: list[tuple[str, int]] = []
        self.foreign_name_choices: list[tuple[str, int]] = []
        self.foreign_names_by_oracle: dict[str, list[str]] = {}
        self.set_booster_max: dict[str, int] = {}
        self.oracle_collectors: dict[str, set[str]] = {}
        self.loaded = False
        self._norm_cache: dict[str, dict[str, tuple[str, int]]] = {}

    def _normalized_choices(self, key: str, choices: list[tuple[str, int]]) -> dict[str, tuple[str, int]]:
        """Map normalized name -> (original name, card index), built once.

        Fuzzy matching against accent-free keys keeps OCR noise (missing
        accents, mojibake) from steering ``extractOne`` toward the wrong card.
        """
        cache = self.__dict__.setdefault("_norm_cache", {})
        result = cache.get(key)
        if result is None:
            result = {}
            for name, index in choices:
                normalized = normalize_text(name)
                if normalized:
                    result.setdefault(normalized, (name, index))
            cache[key] = result
        return result

    def _pt_norm_choices(self) -> dict[str, tuple[str, int]]:
        return self._normalized_choices("pt", self.pt_name_choices)

    def _en_norm_choices(self) -> dict[str, tuple[str, int]]:
        return self._normalized_choices("en", self.name_choices)

    def _face_norm_choices(self) -> dict[str, tuple[str, int]]:
        return self._normalized_choices("face", self.face_name_choices)

    def _foreign_norm_choices(self) -> dict[str, tuple[str, int]]:
        return self._normalized_choices("foreign", self.foreign_name_choices)

    def _build_foreign_names(self) -> None:
        """Printed names of Latin-script foreign printings, mapped to their
        English counterpart. These cover foreign-only sets (FBB, 4BB, REN,
        RIN...) whose cards otherwise have no name in any matching pool."""
        self.foreign_name_choices.clear()
        self.foreign_names_by_oracle.clear()
        en_index_by_oracle: dict[str, int] = {}
        for index, card in enumerate(self.cards):
            if card.get("lang") != "en" or card.get("layout") in _NON_CATALOG_LAYOUTS:
                continue
            oracle_id = card.get("oracle_id") or card.get("name")
            en_index_by_oracle.setdefault(oracle_id, index)
        en_keys = {normalize_text(name) for name, _index in self.name_choices}
        pt_keys = {normalize_text(name) for name, _index in self.pt_name_choices}
        seen = set()
        for card in self.cards:
            if card.get("lang") not in _FOREIGN_NAME_LANGS:
                continue
            if card.get("layout") in _NON_CATALOG_LAYOUTS:
                continue
            oracle_id = card.get("oracle_id") or card.get("name")
            en_index = en_index_by_oracle.get(oracle_id)
            if en_index is None:
                continue
            printed_names = [card.get("printed_name") or ""]
            for face in card.get("card_faces") or []:
                printed_names.append(face.get("printed_name") or "")
            for printed_name in printed_names:
                printed_name = printed_name.strip()
                key = normalize_text(printed_name)
                if not key or key in seen or key in en_keys or key in pt_keys:
                    continue
                seen.add(key)
                self.foreign_name_choices.append((printed_name, en_index))
                self.foreign_names_by_oracle.setdefault(oracle_id, []).append(printed_name)

    def download(self, log) -> None:
        DATA_DIR.mkdir(exist_ok=True)
        log("Buscando metadados de bulk data do Scryfall...")
        bulk = requests.get("https://api.scryfall.com/bulk-data", timeout=60, headers=HTTP_HEADERS)
        bulk.raise_for_status()
        bulk_data = bulk.json()["data"]
        default_cards = next(item for item in bulk_data if item["type"] == "default_cards")

        log("Baixando Default Cards do Scryfall para reconhecer impressões...")
        self._download_file(default_cards["download_uri"], SCRYFALL_CARDS_PATH, log)

        log("Baixando lista de coleções...")
        sets = requests.get("https://api.scryfall.com/sets", timeout=60, headers=HTTP_HEADERS)
        sets.raise_for_status()
        SETS_PATH.write_text(json.dumps(sets.json(), ensure_ascii=False), encoding="utf-8")

        log("Baixando traduções do MTGJSON AtomicCards...")
        self._download_file("https://mtgjson.com/api/v5/AtomicCards.json.gz", MTGJSON_ATOMIC_PATH, log)

        if INDEX_PATH.exists():
            INDEX_PATH.unlink()
        log("Base baixada. Recriando índice local...")
        self.load(force_rebuild=True, log=log)

    @staticmethod
    def _download_file(url: str, path: Path, log) -> None:
        temp_path = path.with_suffix(".tmp")
        with requests.get(url, stream=True, timeout=120, headers=HTTP_HEADERS) as response:
            response.raise_for_status()
            total = int(response.headers.get("content-length", "0") or "0")
            done = 0
            last_log = time.time()
            with temp_path.open("wb") as handle:
                for chunk in response.iter_content(chunk_size=1024 * 1024):
                    if not chunk:
                        continue
                    handle.write(chunk)
                    done += len(chunk)
                    if time.time() - last_log > 2:
                        if total:
                            log(f"Download: {done / total:.0%}")
                        else:
                            log(f"Download: {done // (1024 * 1024)} MB")
                        last_log = time.time()
        temp_path.replace(path)

    def load(self, force_rebuild: bool = False, log=lambda _msg: None) -> None:
        if self.loaded and not force_rebuild:
            return
        self._norm_cache = {}
        if not SCRYFALL_CARDS_PATH.exists():
            raise FileNotFoundError("Base não encontrada. Clique em Atualizar base primeiro.")

        if INDEX_PATH.exists() and not force_rebuild:
            log("Carregando índice local...")
            with INDEX_PATH.open("rb") as handle:
                state = pickle.load(handle)
            if state.get("index_version") != INDEX_VERSION:
                log("Índice antigo detectado. Recriando...")
                return self.load(force_rebuild=True, log=log)
            self.__dict__.update(state)
            self.loaded = True
            return

        log("Lendo arquivo de cartas...")
        with SCRYFALL_CARDS_PATH.open("r", encoding="utf-8") as handle:
            self.cards = json.load(handle)

        if SETS_PATH.exists():
            with SETS_PATH.open("r", encoding="utf-8") as handle:
                sets_payload = json.load(handle)
            self.sets = {item["code"].upper(): item for item in sets_payload.get("data", [])}

        self.by_print.clear()
        self.pt_by_oracle.clear()
        self.en_by_oracle.clear()
        self.mtgjson_pt_by_name.clear()
        self.name_choices.clear()
        self.pt_name_choices.clear()
        self.face_name_choices.clear()
        self.foreign_name_choices.clear()
        self.foreign_names_by_oracle.clear()
        self.set_booster_max.clear()
        self.oracle_collectors.clear()

        seen_choices = set()
        seen_face_choices = set()
        for index, card in enumerate(self.cards):
            set_code = (card.get("set") or "").upper()
            collector = collector_key(card.get("collector_number"))
            if set_code and collector:
                self.by_print.setdefault((set_code, collector), []).append(card)
            oracle_id = card.get("oracle_id") or card.get("name")
            if oracle_id and collector:
                self.oracle_collectors.setdefault(oracle_id, set()).add(collector)
            if card.get("lang") == "en" and card.get("booster") and str(card.get("collector_number", "")).isdigit():
                collector_number = int(card["collector_number"])
                self.set_booster_max[set_code] = max(self.set_booster_max.get(set_code, 0), collector_number)

            oracle_id = card.get("oracle_id") or card.get("name")
            if card.get("lang") == "pt":
                self.pt_by_oracle.setdefault(oracle_id, []).append(card)
            elif card.get("lang") == "en":
                self.en_by_oracle.setdefault(oracle_id, []).append(card)
                # Art-series prints, tokens and emblems reuse a real card's name
                # (art series often duplicates it as "X // X", and there are token
                # "Goblin"/"Soldier" cards) but are not collectible cards anyone
                # scans to catalog. Keeping them out of the fuzzy name/face pools
                # stops the matcher from returning the art print or token instead
                # of the real card (e.g. AMSH art card instead of the MSC card, or
                # a "Goblin" token instead of "Zhur-Taa Goblin").
                if card.get("layout") in _NON_CATALOG_LAYOUTS:
                    continue
                for candidate in [card.get("name"), card.get("printed_name")]:
                    key = normalize_text(candidate or "")
                    if key and key not in seen_choices:
                        self.name_choices.append((candidate, index))
                        seen_choices.add(key)
                for face in card.get("card_faces") or []:
                    face_name = face.get("name") or ""
                    key = normalize_text(face_name)
                    if key and key not in seen_face_choices:
                        self.face_name_choices.append((face_name, index))
                        seen_face_choices.add(key)

        if MTGJSON_ATOMIC_PATH.exists():
            log("Lendo traduções em português...")
            self.mtgjson_pt_by_name = self._load_mtgjson_portuguese()
            for english_name, translation in MANUAL_PT_TRANSLATIONS.items():
                self.mtgjson_pt_by_name[normalize_text(english_name)] = translation
            seen_pt_choices = set()
            for index, card in enumerate(self.cards):
                if card.get("lang") != "en":
                    continue
                if card.get("layout") in _NON_CATALOG_LAYOUTS:
                    continue
                pt_card = self.mtgjson_pt_by_name.get(normalize_text(card.get("name", "")))
                pt_name = (pt_card or {}).get("printed_name", "")
                key = normalize_text(pt_name)
                if key and key not in seen_pt_choices:
                    self.pt_name_choices.append((pt_name, index))
                    seen_pt_choices.add(key)
                    for part in re.split(r"\s*/\s*", pt_name):
                        part = part.strip()
                        part_key = normalize_text(part)
                        if part_key and part_key not in seen_face_choices:
                            self.face_name_choices.append((part, index))
                            seen_face_choices.add(part_key)

        self._build_foreign_names()

        state = {
            "cards": self.cards,
            "sets": self.sets,
            "by_print": self.by_print,
            "pt_by_oracle": self.pt_by_oracle,
            "en_by_oracle": self.en_by_oracle,
            "mtgjson_pt_by_name": self.mtgjson_pt_by_name,
            "name_choices": self.name_choices,
            "pt_name_choices": self.pt_name_choices,
            "face_name_choices": self.face_name_choices,
            "foreign_name_choices": self.foreign_name_choices,
            "foreign_names_by_oracle": self.foreign_names_by_oracle,
            "set_booster_max": self.set_booster_max,
            "oracle_collectors": self.oracle_collectors,
            "loaded": True,
            "index_version": INDEX_VERSION,
        }
        with INDEX_PATH.open("wb") as handle:
            pickle.dump(state, handle, protocol=pickle.HIGHEST_PROTOCOL)
        self.loaded = True

    OCR_SET_FIXES = {
        "LAIS": "M15",
        "LAAIS": "M15",
        "LAA1S": "M15",
        "LAAI5": "M15",
        "L4A15": "M15",
        "M1S": "M15",
        "M1L5": "M15",
        "MOR": "MOM",
        "MRO": "MOM",
        "MOH": "MOM",
        "HOM": "MOM",
        "FRE": "FRF",
        "ECT": "ECL",
        "ECI": "ECL",
        "EC1": "ECL",
        "RN4": "RNA",
        "RMA": "RNA",
        "RNI": "RNA",
        "AFI": "AFR",
        "AFP": "AFR",
        "GAN": "GTC",
        "GRMN": "GTC",
        "GRAN": "GTC",
        "GRMN": "GTC",
        "ORM": "GTC",
        "FOE": "EOE",
        "EOF": "EOE",
        "FON": "FDN",
        "FOK": "FDN",
        "OMU": "DMU",
        "OPM": "DMU",
        "NEOS": "NEO",
        "MEO": "NEO",
        "SEO": "NEO",
        "S82": "SS2",
        "SB2": "SS2",
        "3B2": "SS2",
        "382": "SS2",
    }

    def _known_set_codes(self) -> set[str]:
        return {set_code for set_code, _collector in self.by_print.keys()}

    def _sets_for_card_count(self, total: int) -> list[str]:
        matches = []
        seen = set()
        for code, info in self.sets.items():
            if info.get("card_count") == total and code not in seen:
                matches.append(code)
                seen.add(code)
        for code, booster_max in self.set_booster_max.items():
            if booster_max == total and code not in seen:
                matches.append(code)
                seen.add(code)
        return matches

    @staticmethod
    def _is_likely_collector_fraction(collector: str, total: int) -> bool:
        if total < 30:
            return False
        collector_digits = re.sub(r"\D", "", collector)
        if not collector_digits:
            return False
        return int(collector_digits) <= total

    @staticmethod
    def _collector_fractions_in_text(raw_text: str) -> list[tuple[str, int]]:
        fractions = []
        text = raw_text.upper()
        for match in re.finditer(r"\b([0-9OIL]{1,4}[A-Z]?)\s*/\s*([0-9OILSE]{1,5})\b", text):
            collector = ocr_collector_key(match.group(1))
            total_raw = match.group(2).translate(str.maketrans({"O": "0", "I": "1", "L": "1", "E": "8", "S": "5"}))
            if re.fullmatch(r"\d{1,5}", total_raw):
                total = int(total_raw)
                if ScryfallDatabase._is_likely_collector_fraction(collector, total):
                    fractions.append((collector, total))
        return fractions

    def _corrected_set(
        self,
        token: str,
        known_sets: set[str],
        collector: str = "",
        total: int = 0,
    ) -> str:
        fixed = self.OCR_SET_FIXES.get(token, "")
        if token in known_sets:
            if (
                fixed
                and fixed in known_sets
                and collector
                and not self.by_print.get((token, collector))
                and self.by_print.get((fixed, collector))
            ):
                return fixed
            return token
        if fixed and fixed in known_sets:
            return fixed
        if total and collector:
            matching_sets = self._sets_for_card_count(total)
            if len(matching_sets) == 1 and self.by_print.get((matching_sets[0], collector)):
                return matching_sets[0]
        if fuzz and process and len(token) >= 2:
            candidates = list(known_sets)
            if total and collector:
                count_sets = self._sets_for_card_count(total)
                if count_sets:
                    candidates = count_sets
            match = process.extractOne(token, candidates, scorer=fuzz.ratio)
            min_score = 86 if len(token) <= 3 else 80
            if match and match[1] >= min_score:
                if self.by_print.get((match[0], collector)):
                    return match[0]
        return ""

    def _small_set_collector_variants(self, set_code: str, collector: str) -> list[str]:
        set_info = self.sets.get(set_code) or {}
        set_total = int(set_info.get("card_count") or self.set_booster_max.get(set_code) or 0)
        if not (1 <= set_total <= 20):
            return []
        digits = re.sub(r"\D", "", collector_key(collector))
        if len(digits) <= 1:
            return []
        variants = []
        for size in (1, 2):
            suffix = digits[-size:]
            key = collector_key(suffix)
            if key in variants:
                continue
            if re.fullmatch(r"\d+", key) and int(key) <= set_total and self.by_print.get((set_code, key)):
                variants.append(key)
        return variants

    @staticmethod
    def _collector_token_variants(token: str) -> list[str]:
        variants = []

        def add(value: str) -> None:
            key = ocr_collector_key(value)
            if key and key not in variants:
                variants.append(key)

        if re.fullmatch(r"0*\d{1,4}[A-Z]?", token):
            add(token)
        if re.fullmatch(r"S\d{1,3}[A-Z]?", token):
            add(f"3{token[1:]}")
        if re.fullmatch(r"S[O0]\d{1,2}[A-Z]?", token):
            add(f"3{token[1:].translate(str.maketrans({'O': '0'}))}")
        return variants

    def _print_candidates(self, set_code: str, collector: str) -> list[dict]:
        candidates = [card for card in self.by_print.get((set_code, collector), []) if card.get("lang") == "en"]
        return candidates or self.by_print.get((set_code, collector), [])

    def _pick_best_print_candidate(self, candidates: list[dict], raw_text: str) -> dict | None:
        if not candidates:
            return None
        if len(candidates) == 1:
            return candidates[0]
        scored = []
        for card in candidates:
            score = 0
            if self._card_name_in_raw_text(card, raw_text):
                score += 100
            pt_card = self.portuguese_for(card)
            pt_name = (pt_card or {}).get("printed_name") or ""
            if pt_name and normalize_text(pt_name) in normalize_text(raw_text):
                score += 80
            scored.append((score, card))
        scored.sort(key=lambda item: item[0], reverse=True)
        if scored[0][0] > 0:
            return scored[0][1]
        return candidates[0]

    def _set_codes_in_text(self, raw_text: str) -> set[str]:
        known_sets = self._known_set_codes()
        tokens = re.findall(r"[A-Z0-9]+", raw_text.upper())
        found = set()
        for index, token in enumerate(tokens):
            if token in known_sets and self._set_token_has_print_context(tokens, index):
                found.add(token)
                continue
            fixed = self.OCR_SET_FIXES.get(token, "")
            if fixed and fixed in known_sets and self._set_fix_has_print_context(token, tokens, index):
                found.add(fixed)
        if fuzz and process:
            for index, token in enumerate(tokens):
                if len(token) != 3 or token in found or not self._set_token_has_print_context(tokens, index):
                    continue
                match = process.extractOne(token, list(known_sets), scorer=fuzz.ratio)
                if match and match[1] >= 88:
                    found.add(match[0])
        found.update(self._symbol_set_codes_in_text(raw_text, known_sets))
        return found

    def _is_guessed_set_code(self, set_code: str, token: str) -> bool:
        """True when the code came from fuzzy correction rather than a literal
        or curated-fix read of the token."""
        return set_code != token and self.OCR_SET_FIXES.get(token, "") != set_code

    @staticmethod
    def _footer_shape_context(tokens: list[str], index: int) -> bool:
        """True footer shape around a set-code token: a language code on the
        right ("WOE PT") or a zero-padded collector ("0178") on the left. A
        stray number from the copyright strip ("109%", "1095") never has the
        leading zero."""
        left_window = tokens[max(0, index - 4) : index]
        right_window = tokens[index + 1 : index + 5]
        if any(value in {"EN", "PT", "ES", "FR", "DE", "IT", "JP", "KO"} for value in right_window):
            return True
        return any(re.fullmatch(r"0\d{2,3}[A-Z]?", value) for value in left_window)

    @staticmethod
    def _set_token_has_print_context(tokens: list[str], index: int) -> bool:
        token = tokens[index]
        left_window = tokens[max(0, index - 4) : index]
        if token not in AMBIGUOUS_SET_CODE_WORDS and not (
            len(token) == 3 and token.isalpha()
        ):
            return True
        # A word that shows up in normal rules text ("one", "war", "the") is
        # only a set code in true footer shape.
        if token in AMBIGUOUS_SET_CODE_WORDS:
            return ScryfallDatabase._footer_shape_context(tokens, index)
        if ScryfallDatabase._footer_shape_context(tokens, index):
            return True
        for value in left_window:
            number_match = re.fullmatch(r"0*(\d{3,4})[A-Z]?", value)
            if number_match and int(number_match.group(1)) <= 999:
                return True
        return False

    @classmethod
    def _set_fix_has_print_context(cls, token: str, tokens: list[str], index: int) -> bool:
        if not (len(token) == 3 and token.isalpha()):
            return True
        return cls._set_token_has_print_context(tokens, index)

    @staticmethod
    def _symbol_set_codes_in_text(raw_text: str, known_sets: set[str]) -> set[str]:
        found = set()
        for line in raw_text.splitlines():
            normalized = normalize_text(line)
            if not re.search(
                r"\b(criatura|creature|instantanea|instant|feiticaria|sorcery|"
                r"encantamento|enchantment|artefato|artifact|terreno|land)\b",
                normalized,
            ):
                continue
            tokens = re.findall(r"[A-Z0-9]+", line.upper())
            if tokens and tokens[-1] in {"X", "XX"} and "10E" in known_sets:
                found.add("10E")
        return found

    def _namebar_candidates(self, namebar_text: str) -> list[str]:
        candidates = []
        for line in ranked_namebar_lines(namebar_text):
            for value in (line, normalize_ocr_name(line)):
                if value and value not in candidates:
                    candidates.append(value)
        normalized_all = normalize_ocr_name(namebar_text)
        if normalized_all and normalized_all not in candidates and is_plausible_namebar(normalized_all):
            candidates.append(normalized_all)
        return candidates

    def _foreign_names_for(self, card: dict) -> list[str]:
        oracle_id = card.get("oracle_id") or card.get("name")
        names = []
        for printed_name in self.foreign_names_by_oracle.get(oracle_id, []):
            names.append(printed_name)
            names.extend(split_face_values(printed_name))
        return names

    def _best_name_coverage(self, card: dict, line: str) -> int:
        """Best coverage of any of the card's names against a namebar line."""
        normalized_line = normalize_text(line)
        names = [card.get("name", "")]
        for face in card.get("card_faces") or []:
            if face.get("name"):
                names.append(face["name"])
        pt_card = self.portuguese_for(card)
        if pt_card:
            printed = pt_card.get("printed_name") or pt_card.get("name") or ""
            names.append(printed)
            names.extend(part.strip() for part in re.split(r"\s*/\s*", printed))
        names.extend(self._foreign_names_for(card))
        return max((self._hint_word_coverage(name, normalized_line) for name in names if name), default=0)

    def _fuzzy_namebar(self, namebar_text: str) -> dict | None:
        # Evaluate every plausible namebar line and keep the match whose name is
        # most fully present in the line it came from. This stops a short name
        # (e.g. the plane "Naya") from winning over a fuller one ("Medalhão de
        # Naya") just because a noisier line was scored higher.
        best_card = None
        best_score = (-1, -1.0, -1)

        def add_option(options: list[tuple[dict, float]], seen: set[str], card: dict | None, score: float) -> None:
            if not card:
                return
            key = card.get("oracle_id") or card.get("id") or card.get("name")
            if key in seen:
                return
            seen.add(key)
            options.append((card, score))

        def candidate_options(candidate: str) -> list[tuple[dict, float]]:
            options: list[tuple[dict, float]] = []
            seen: set[str] = set()
            strict = self._fuzzy_name(candidate, strict_words=True, min_score=84, prefer_en=True)
            add_option(options, seen, strict, 100.0)
            if not strict and is_plausible_namebar(candidate):
                relaxed = self._fuzzy_name(candidate, strict_words=False, min_score=76, prefer_en=True)
                add_option(options, seen, relaxed, 90.0)
            if not process or not fuzz:
                return options

            normalized = normalize_text(candidate)
            if not normalized:
                return options
            # WRatio overvalues short substring cards in noisy title bars
            # ("Seeker" inside "Mindseeker"). Plain ratio keeps fuller names in
            # the candidate pool, then word coverage chooses the title that
            # explains the most OCR words.
            for choices in (
                self._en_norm_choices(),
                self._pt_norm_choices(),
                self._foreign_norm_choices(),
            ):
                if not choices:
                    continue
                for match in process.extract(normalized, choices.keys(), scorer=fuzz.ratio, limit=12):
                    if match[1] < 74:
                        continue
                    _matched_name, card_index = choices[match[0]]
                    add_option(options, seen, self.cards[card_index], float(match[1]))
            return options

        for candidate in self._namebar_candidates(namebar_text):
            candidate_words = [word for word in normalize_text(candidate).split() if len(word) >= 3]
            for card, match_score in candidate_options(candidate):
                cov = self._best_name_coverage(card, candidate)
                if cov <= 0:
                    continue
                card_words = [word for word in normalize_text(card.get("name", "")).split() if len(word) >= 3]
                if len(candidate_words) == 1 and len(card_words) > 1:
                    continue
                if len(candidate_words) >= 2 and cov < 2:
                    continue
                score = (cov, match_score, min(len(card_words), 6))
                if score > best_score:
                    best_score = score
                    best_card = card
        return best_card

    def _fuzzy_foreign_namebar(self, namebar_text: str) -> tuple[dict | None, str]:
        """Match the title crop against foreign printed names (FBB French,
        4BB Spanish...). Old foreign cards have no footer, so a fuzzy namebar
        read anchored by one solid name word ("dUrza" -> "Urza") is the only
        signal available."""
        if not namebar_text or not process or not fuzz:
            return None, ""
        choices = self._foreign_norm_choices()
        if not choices:
            return None, ""
        best_card = None
        best_name = ""
        best_score = 0.0
        for candidate in self._namebar_candidates(namebar_text):
            normalized = normalize_text(candidate)
            candidate_words = [word for word in normalized.split() if len(word) >= 3]
            if not candidate_words or len(candidate_words) > 5:
                continue
            compact = re.sub(r"\s+", "", normalized)
            if len(compact) < 8:
                continue
            for match in process.extract(normalized, choices.keys(), scorer=fuzz.ratio, limit=5):
                if match[1] < 82:
                    continue
                original_name, card_index = choices[match[0]]
                name_words = [word for word in match[0].split() if len(word) >= 4]
                if not name_words:
                    continue
                compact_name = re.sub(r"\s+", "", match[0])
                if fuzz.ratio(compact, compact_name) < 84:
                    continue
                if not any(
                    words_fuzzy_match(name_word, word, 85)
                    for name_word in name_words
                    for word in candidate_words
                ):
                    continue
                if float(match[1]) > best_score:
                    best_score = float(match[1])
                    best_card = self.cards[card_index]
                    best_name = original_name
        return best_card, best_name

    def _find_by_face_names_in_text(self, raw_text: str) -> tuple[dict | None, str]:
        if not raw_text or not self.face_name_choices or not process or not fuzz:
            return None, ""
        choices = self._face_norm_choices()
        best_card = None
        best_name = ""
        best_score = 0
        for line in raw_text.splitlines():
            cleaned = clean_ocr_line(line)
            normalized = normalize_text(cleaned)
            words = normalized.split()
            if not words or len(words) > 4:
                continue
            if not any(len(word) >= 4 for word in words):
                continue
            for size in range(min(3, len(words)), 0, -1):
                for start in range(0, len(words) - size + 1):
                    candidate = " ".join(words[start : start + size])
                    if len(candidate) < 5:
                        continue
                    if candidate in COMMON_EN_OCR_WORDS or candidate in COMMON_PT_OCR_WORDS:
                        continue
                    if candidate in MTG_KEYWORD_NAME_STOPWORDS:
                        continue
                    match = process.extractOne(candidate, choices.keys(), scorer=fuzz.WRatio)
                    if not match:
                        continue
                    original_name, card_index = choices[match[0]]
                    matched_len = len(match[0])
                    min_score = 96 if len(candidate) < 7 else 90
                    if match[1] < min_score:
                        continue
                    if len(candidate) < matched_len * 0.55:
                        continue
                    matched_words = match[0].split()
                    if len(matched_words) == 1 and candidate != match[0]:
                        continue
                    if match[1] > best_score:
                        best_score = match[1]
                        best_name = original_name
                        best_card = self.cards[card_index]
        if best_card:
            return best_card, best_name
        return None, ""

    def _fraction_supports_print(self, set_code: str, collector: str, fractions: list[tuple[str, int]]) -> int:
        score = 0
        for fraction_collector, total in fractions:
            if fraction_collector != collector:
                continue
            if set_code in self._sets_for_card_count(total):
                score = max(score, 150)
        return score

    def _distinctive_name_in_raw_text(self, card: dict, raw_text: str) -> bool:
        names = [card.get("name", "")]
        for face in card.get("card_faces") or []:
            face_name = face.get("name") or ""
            if face_name:
                names.append(face_name)
        pt_card = self.portuguese_for(card)
        if pt_card:
            printed_name = pt_card.get("printed_name") or pt_card.get("name") or ""
            names.append(printed_name)
            for part in re.split(r"\s*/\s*", printed_name):
                part = part.strip()
                if part:
                    names.append(part)
        names.extend(self._foreign_names_for(card))
        normalized_raw = normalize_text(raw_text)
        raw_words = [word for word in normalized_raw.split() if len(word) >= 3]
        for card_name in names:
            if not card_name:
                continue
            normalized_name = normalize_text(card_name)
            if normalized_phrase_in_text(normalized_name, normalized_raw):
                return True
            name_words = [
                word for word in normalized_name.split() if len(word) >= 4 and word not in COMMON_PT_OCR_WORDS
            ]
            if not name_words:
                name_words = [word for word in normalized_name.split() if len(word) >= 4]
            if not name_words:
                continue
            if all(word in COMMON_PT_OCR_WORDS for word in name_words):
                continue
            required = 2 if len(name_words) >= 2 else 1
            if fuzzy_word_count(name_words, raw_words) >= required:
                return True
            if fuzz:
                for line in raw_text.splitlines():
                    cleaned = clean_ocr_line(line)
                    if len(cleaned.split()) > 6:
                        continue
                    normalized_line = normalize_ocr_name(cleaned)
                    line_words = [word for word in normalized_line.split() if len(word) >= 3]
                    if name_words and len(name_words) == 1:
                        word = name_words[0]
                        if not any(
                            words_fuzzy_match(word, raw_word, 88) and len(raw_word) >= max(5, len(word) * 0.55)
                            for raw_word in line_words
                        ):
                            continue
                    if fuzz.WRatio(card_name, cleaned) >= 84 or fuzz.WRatio(card_name, normalized_line) >= 76:
                        return True
        return False

    def _print_evidence_in_ocr(self, card: dict, ctx: OcrMatchContext) -> bool:
        ocr = ctx.ocr
        set_code = (card.get("set") or "").upper()
        collector = collector_key(card.get("collector_number"))
        known_sets = self._known_set_codes()
        if ocr.set_code and ocr.collector_number:
            if ocr.set_code.upper() == set_code and collector_key(ocr.collector_number) == collector:
                return True
        for index, token in enumerate(ctx.tokens):
            if collector not in self._collector_token_variants(token):
                continue
            window = {
                self._corrected_set(value, known_sets, collector) or value
                for value in ctx.tokens[index + 1 : index + 16]
            }
            if set_code in window:
                return True
        fraction_collectors = {fraction_collector for fraction_collector, _total in ctx.fractions}
        if set_code in ctx.sets_in_text and collector in fraction_collectors:
            return True
        if collector in self._collector_numbers_in_text(ocr.raw_text):
            if set_code in ctx.sets_in_text or set_code in ctx.tokens:
                return True
        return False

    def _print_pair_occurrences(self, card: dict, ctx: OcrMatchContext) -> int:
        set_code = (card.get("set") or "").upper()
        collector = collector_key(card.get("collector_number"))
        known_sets = self._known_set_codes()
        count = 0
        for index, token in enumerate(ctx.tokens):
            if collector not in self._collector_token_variants(token):
                continue
            window = {
                self._corrected_set(value, known_sets, collector) or value
                for value in ctx.tokens[index + 1 : index + 16]
            }
            if set_code in window:
                count += 1
        return count

    def _reliable_hint_card(self, ocr: OcrResult) -> dict | None:
        # Prefer the name bar (the card title) over face names: a single common
        # word in the body (e.g. the creature type "Goblin") can match a split
        # card's face ("Weird // Goblin") and give a misleading hint.

        namebar_pt, namebar_pt_name = self._fuzzy_portuguese_text(ocr.namebar_text, namebar_mode=True)
        if namebar_pt and (
            self._distinctive_name_in_raw_text(namebar_pt, ocr.raw_text)
            or self._best_name_coverage(namebar_pt, normalize_text(namebar_pt_name)) >= 1
        ):
            return namebar_pt
        namebar_en = self._fuzzy_namebar(ocr.namebar_text)
        if namebar_en and self._distinctive_name_in_raw_text(namebar_en, ocr.raw_text):
            return namebar_en
        card, _matched_line = self._fuzzy_english_title_in_text(ocr.raw_text)
        if card and self._distinctive_name_in_raw_text(card, ocr.raw_text):
            return card

        pt_words_card, pt_words_name = self._fuzzy_portuguese_words_in_text(ocr.raw_text)
        if (
            pt_words_card
            and len(normalize_text(pt_words_name).split()) >= 2
            and self._distinctive_name_in_raw_text(pt_words_card, ocr.raw_text)
        ):
            return pt_words_card
        face_card, _ = self._find_by_face_names_in_text(ocr.raw_text)
        if face_card and self._distinctive_name_in_raw_text(face_card, ocr.raw_text):
            return face_card
        if pt_words_card and self._distinctive_name_in_raw_text(pt_words_card, ocr.raw_text):
            return pt_words_card
        if ocr.name_hint and self._likely_name_hint(ocr.name_hint):
            normalized_hint = normalize_ocr_name(ocr.name_hint)
            if is_plausible_namebar(normalized_hint):
                card = self._fuzzy_name(normalized_hint, strict_words=False, min_score=76)
                if (
                    card
                    and self._hint_supports_card(card, ocr.name_hint)
                    and self._distinctive_name_in_raw_text(card, ocr.raw_text)
                ):
                    return card
            if is_plausible_namebar(ocr.name_hint):
                card = self._fuzzy_name(ocr.name_hint, strict_words=True, min_score=84)
                if (
                    card
                    and self._hint_supports_card(card, ocr.name_hint)
                    and self._distinctive_name_in_raw_text(card, ocr.raw_text)
                ):
                    return card
        return None

    @staticmethod
    def _hint_supports_card(card: dict, hint: str) -> bool:
        """A one-word card name matched out of a long hint line is a phrase
        hit (flavor "a bencao do Divino" -> "Bênção"), not a title read."""
        hint_words = [word for word in normalize_text(hint).split() if len(word) >= 3]
        name_words = normalize_text(card.get("name", "")).split()
        return len(name_words) > 1 or len(hint_words) <= 3

    def _fuzzy_portuguese_words_in_text(self, raw_text: str) -> tuple[dict | None, str]:
        if not raw_text or not self.pt_name_choices:
            return None, ""
        normalized_raw = normalize_text(raw_text)
        raw_words = [word for word in normalized_raw.split() if len(word) >= 4]
        if len(raw_words) < 2:
            return None, ""
        normalized_lines = [
            (normalize_text(line), [word for word in normalize_text(line).split() if len(word) >= 3])
            for line in raw_text.splitlines()
        ]
        # Words on short, title-like lines. A name whose evidence is a single
        # word must come from one of these: an exact token inside a longer
        # line (copyright garbage "... bo sabia a rine", or the flavor line
        # "A Gloria surgiu de dentro dela") is not name evidence.
        short_line_words = [
            word
            for _normalized_line, line_words in normalized_lines
            if line_words and len(line_words) <= 3
            for word in line_words
            if len(word) >= 4
        ]
        best_word_card = None
        best_word_name = ""
        best_word_score = 0
        for pt_name, index in self.pt_name_choices:
            normalized_pt_name = normalize_text(pt_name)
            pt_words = [word for word in normalized_pt_name.split() if len(word) >= 4]
            if not pt_words:
                continue
            distinctive_words = [word for word in pt_words if word not in COMMON_PT_OCR_WORDS]
            match_words = distinctive_words or pt_words
            all_pt_words = [word for word in normalized_pt_name.split() if len(word) >= 3]
            first_word = all_pt_words[0] if all_pt_words else ""
            if len(all_pt_words) > 1 and first_word:
                if not any(words_fuzzy_match(first_word, word, 86) for word in raw_words):
                    continue
            # A single distinctive word is weak evidence: at 86 an OCR noise
            # token passes ("cone" for "Clone", English "controler" for
            # "Controlar"), and at 90 "pagar" still reads as "Apagar". Require
            # a near-exact read on a title-like line, and at least 5 letters —
            # 4-letter tokens ("alem") show up verbatim in OCR garbage.
            if len(match_words) == 1 and (
                len(match_words[0]) < 5
                or not fuzzy_word_count(match_words, short_line_words, threshold=92)
            ):
                continue
            matches = fuzzy_word_count(match_words, raw_words)
            required = min(3, len(match_words)) if len(match_words) >= 3 else min(2, len(match_words))
            if matches < required:
                continue
            line_bonus = 0
            if len(all_pt_words) >= 2:
                for line_index, (normalized_line, line_words) in enumerate(normalized_lines):
                    if not normalized_line:
                        continue
                    if normalized_pt_name and normalized_pt_name in normalized_line:
                        line_bonus = max(line_bonus, 900 - min(line_index, 80))
                        continue
                    if fuzzy_word_count(all_pt_words, line_words, threshold=88) >= len(all_pt_words):
                        line_bonus = max(line_bonus, 650 - min(line_index, 80))
            score = line_bonus + matches * 100 + min(len(pt_words), 6)
            if score > best_word_score:
                best_word_score = score
                best_word_name = pt_name
                best_word_card = self.cards[index]
        if best_word_card:
            return best_word_card, best_word_name
        return None, ""

    def _oracle_has_collector(self, oracle_id: str, collector: str) -> bool:
        return collector in self.oracle_collectors.get(oracle_id, set())

    def _manual_portuguese_title_in_text(self, raw_text: str) -> tuple[dict | None, str]:
        if not raw_text:
            return None, ""
        import difflib

        manual_names = {
            normalize_text(data.get("printed_name", "")): english_name
            for english_name, data in MANUAL_PT_TRANSLATIONS.items()
            if data.get("printed_name")
        }
        manual_names = {key: value for key, value in manual_names.items() if key}
        if not manual_names:
            return None, ""
        normalized_raw = normalize_text(raw_text)
        compact_raw = re.sub(r"\s+", "", normalized_raw)
        raw_words = [word for word in normalized_raw.split() if len(word) >= 3]
        best_card = None
        best_name = ""
        best_score = 0.0
        for line in raw_text.splitlines():
            cleaned = clean_ocr_line(line)
            words = title_words(cleaned)
            if not words or len(words) > 4:
                continue
            candidate = words[0]
            if candidate in COMMON_PT_OCR_WORDS or candidate in COMMON_EN_OCR_WORDS:
                continue
            if process and fuzz:
                match = process.extractOne(candidate, manual_names.keys(), scorer=fuzz.ratio)
            else:
                matched_name = ""
                matched_score = 0.0
                for manual_name in manual_names:
                    score = difflib.SequenceMatcher(None, candidate, manual_name).ratio() * 100
                    if score > matched_score:
                        matched_name = manual_name
                        matched_score = score
                match = (matched_name, matched_score, None) if matched_name else None
            if not match:
                continue
            matched_name = match[0]
            if match[1] < 72:
                continue
            english_name = manual_names[matched_name]
            card = self._fuzzy_name(english_name, strict_words=True, min_score=99, prefer_en=True)
            if not card:
                continue
            same_oracle_cards = [
                item
                for item in self.cards
                if item.get("lang") == "en" and item.get("oracle_id") == card.get("oracle_id")
            ] or [card]
            artist_ok = any(
                self._artist_name_in_raw_text(item, normalized_raw, compact_raw, raw_words)
                for item in same_oracle_cards
            )
            support = self._best_oracle_support(card, raw_text)
            if not artist_ok and support < 2:
                continue
            if candidate[:1] != matched_name[:1] and not (artist_ok or support >= 2):
                continue
            score = float(match[1]) + support * 20 + (25 if artist_ok else 0)
            if score > best_score:
                best_score = score
                best_card = card
                best_name = MANUAL_PT_TRANSLATIONS[english_name]["printed_name"]
        if best_card:
            return best_card, best_name
        return None, ""

    def _build_match_context(self, ocr: OcrResult) -> OcrMatchContext:
        fractions = self._collector_fractions_in_text(ocr.raw_text)
        tokens = re.findall(r"[A-Z0-9]+", ocr.raw_text.upper())
        sets_in_text = self._set_codes_in_text(ocr.raw_text)
        namebar_en = self._fuzzy_namebar(ocr.namebar_text)
        namebar_pt, _ = self._fuzzy_portuguese_text(ocr.namebar_text, namebar_mode=True)
        face_card, _ = self._find_by_face_names_in_text(ocr.raw_text)
        pt_words_card, _ = self._fuzzy_portuguese_words_in_text(ocr.raw_text)
        hint_card = self._reliable_hint_card(ocr)
        hint_oracle = hint_card.get("oracle_id") if hint_card else None
        hint_from_title = bool(
            hint_oracle
            and (
                (namebar_en and namebar_en.get("oracle_id") == hint_oracle)
                or (namebar_pt and namebar_pt.get("oracle_id") == hint_oracle)
            )
        )
        return OcrMatchContext(
            ocr=ocr,
            fractions=fractions,
            tokens=tokens,
            sets_in_text=sets_in_text,
            hint_card=hint_card,
            namebar_resolved=bool(
                (namebar_en and self._card_name_in_raw_text(namebar_en, ocr.raw_text))
                or (namebar_pt and self._card_name_in_raw_text(namebar_pt, ocr.raw_text))
                or (face_card and self._card_name_in_raw_text(face_card, ocr.raw_text))
                or pt_words_card
            ),
            hint_from_title=hint_from_title,
        )

    def _find_by_collector_fraction(self, raw_text: str, ctx: OcrMatchContext | None = None) -> tuple[dict | None, str, str]:
        fractions = ctx.fractions if ctx else self._collector_fractions_in_text(raw_text)
        if not fractions:
            return None, "", ""
        sets_in_text = ctx.sets_in_text if ctx else self._set_codes_in_text(raw_text)
        collectors_in_fractions = {collector for collector, _total in fractions}
        if sets_in_text and collectors_in_fractions:
            best_card = None
            best_set = ""
            best_collector = ""
            best_score = -1
            for set_code in sets_in_text:
                for collector in collectors_in_fractions:
                    candidates = self._print_candidates(set_code, collector)
                    if not candidates:
                        continue
                    card = self._pick_best_print_candidate(candidates, raw_text)
                    if not card:
                        continue
                    score = 250
                    score += self._fraction_supports_print(set_code, collector, fractions)
                    if self._card_name_in_raw_text(card, raw_text):
                        score += 100
                    if score > best_score:
                        best_score = score
                        best_card = card
                        best_set = set_code
                        best_collector = collector
            if best_card:
                return best_card, best_set, best_collector

        fraction_counts: dict[tuple[str, int], int] = {}
        for collector, total in fractions:
            fraction_counts[(collector, total)] = fraction_counts.get((collector, total), 0) + 1
        ranked_fractions = sorted(fraction_counts.items(), key=lambda item: item[1], reverse=True)
        for (collector, total), _count in ranked_fractions:
            matching_sets = self._sets_for_card_count(total)
            if not matching_sets:
                continue
            best_card = None
            best_set = ""
            best_score = -1
            for set_code in matching_sets:
                candidates = self._print_candidates(set_code, collector)
                if not candidates:
                    continue
                card = self._pick_best_print_candidate(candidates, raw_text)
                if not card:
                    continue
                score = 50
                score += self._fraction_supports_print(set_code, collector, fractions)
                if set_code in sets_in_text:
                    score += 200
                if self._card_name_in_raw_text(card, raw_text):
                    score += 100
                if score > best_score:
                    best_score = score
                    best_card = card
                    best_set = set_code
            if best_card:
                if len(matching_sets) > 1 and best_score < 150:
                    continue
                return best_card, best_set, collector
        return None, "", ""

    def _find_by_fraction_and_name(
        self, raw_text: str, ctx: OcrMatchContext
    ) -> tuple[dict | None, str, str]:
        """Strongest possible signal: a collector fraction (e.g. "198/249")
        whose printing also has its name in the OCR text. The footer number and
        the card name agree, so this beats every fuzzy path — and it is robust
        to a mangled set code (``IMA`` read as ``IM AS``) because the set is
        derived from the card-count total instead of the OCR'd letters."""
        for collector, total in ctx.fractions:
            candidate_sets = set(ctx.sets_in_text) | set(self._sets_for_card_count(total))
            for set_code in candidate_sets:
                for card in self._print_candidates(set_code, collector):
                    if self._strong_card_name_in_raw_text(card, raw_text):
                        return card, set_code, collector
        return None, "", ""

    def find(self, ocr: OcrResult) -> tuple[dict, str]:
        self.load()
        ctx = self._build_match_context(ocr)

        def finalize(card: dict, reason: str) -> tuple[dict, str]:
            return self._prefer_print_from_ocr(card, ocr.raw_text), reason

        def try_match(card: dict | None, reason: str) -> tuple[dict, str] | None:
            if not card:
                return None
            card = self._prefer_print_from_ocr(card, ocr.raw_text)
            if self.is_confident_match(card, ctx):
                return card, reason
            return None

        def strong_portuguese_namebar_match(card: dict, matched_name: str) -> bool:
            coverage = self._best_name_coverage(card, normalize_text(matched_name))
            significant_words = [
                word for word in normalize_text(matched_name).split() if len(word) >= 3
            ]
            if coverage >= min(2, len(significant_words)):
                return True
            normalized_raw = normalize_text(ocr.raw_text)
            artist_ok = self._artist_name_in_raw_text(
                card,
                normalized_raw,
                re.sub(r"\s+", "", normalized_raw),
                [word for word in normalized_raw.split() if len(word) >= 3],
            )
            return coverage >= 1 and artist_ok and self._best_oracle_support(card, ocr.raw_text) >= 4

        # Identify the card from the dedicated title crop before using isolated
        # footer tokens. Footer evidence still selects the printing afterwards.
        card, matched_line = self._exact_english_title_in_text(ocr.namebar_text)
        matched = try_match(card, f"título OCR '{matched_line}'")
        if matched:
            return matched

        # A collector fraction agreeing with a strongly-present name outranks
        # every fuzzy title path (it also protects against namebar noise that
        # happens to read like a short foreign card name, e.g. "SEIS").
        card, set_code, collector_number = self._find_by_fraction_and_name(ocr.raw_text, ctx)
        if card:
            return finalize(card, f"fração+nome {set_code} #{collector_number}")

        namebar_pt, namebar_pt_name = self._fuzzy_portuguese_text(
            ocr.namebar_text, namebar_mode=True
        )
        if namebar_pt and strong_portuguese_namebar_match(namebar_pt, namebar_pt_name):
            matched = try_match(namebar_pt, f"nome superior PT OCR '{namebar_pt_name}'")
            if matched:
                return matched

        card, matched_line = self._fuzzy_english_title_in_text(ocr.namebar_text)
        matched = try_match(card, f"título superior OCR aproximado '{matched_line}'")
        if matched:
            return matched

        foreign_card, foreign_name = self._fuzzy_foreign_namebar(ocr.namebar_text)
        if foreign_card:
            matched = try_match(foreign_card, f"nome estrangeiro OCR '{foreign_name}'")
            if matched:
                return matched

        if ocr.set_code and ocr.collector_number:
            candidates = self._print_candidates(ocr.set_code.upper(), collector_key(ocr.collector_number))
            if candidates:
                card = self._pick_best_print_candidate(candidates, ocr.raw_text)
                matched = try_match(card, f"código {ocr.set_code.upper()} #{ocr.collector_number}")
                if matched:
                    return matched

        card, matched_line = self._exact_english_self_reference_in_text(ocr.raw_text)
        matched = try_match(card, f"auto-referencia OCR '{matched_line}'")
        if matched:
            return matched

        card, set_code, collector_number = self._find_by_collector_fraction(ocr.raw_text, ctx)
        matched = try_match(card, f"fração collector {set_code} #{collector_number}")
        if matched:
            return matched

        card, set_code, collector_number = self._find_print_in_ocr_text(ocr.raw_text)
        matched = try_match(card, f"rodapé OCR {set_code} #{collector_number}")
        if matched:
            return matched

        card, matched_line = self._exact_english_self_reference_in_text(ocr.raw_text)
        matched = try_match(card, f"auto-referencia OCR '{matched_line}'")
        if matched:
            return matched

        card, matched_line = self._exact_english_title_in_text(ocr.raw_text)
        if card and self._best_oracle_support(card, ocr.raw_text) < 5:
            card = None
        matched = try_match(card, f"título OCR bruto '{matched_line}'")
        if matched:
            return matched

        card, matched_line = self._fuzzy_english_title_in_text(ocr.raw_text)
        matched = try_match(card, f"título OCR aproximado '{matched_line}'")
        if matched:
            return matched

        card, matched_line = self._exact_english_title_in_text(ocr.raw_text)
        if card and self._best_oracle_support(card, ocr.raw_text) < 3:
            card = None
        matched = try_match(card, f"título OCR bruto '{matched_line}'")
        if matched:
            return matched

        manual_pt_card, manual_pt_name = self._manual_portuguese_title_in_text(ocr.raw_text)
        if manual_pt_card:
            return finalize(manual_pt_card, f"nome PT manual OCR '{manual_pt_name}'")

        pt_words_card, pt_words_name = self._fuzzy_portuguese_words_in_text(ocr.raw_text)
        if (
            pt_words_card
            and len(normalize_text(pt_words_name).split()) >= 2
            and self._distinctive_name_in_raw_text(pt_words_card, ocr.raw_text)
        ):
            matched = try_match(pt_words_card, f"nome PT palavras '{pt_words_name}'")
            if matched:
                return matched

        face_card, face_name = self._find_by_face_names_in_text(ocr.raw_text)
        if face_card and self._distinctive_name_in_raw_text(face_card, ocr.raw_text):
            matched = try_match(face_card, f"face OCR '{face_name}'")
            if matched:
                return matched

        if pt_words_card and self._distinctive_name_in_raw_text(pt_words_card, ocr.raw_text):
            matched = try_match(pt_words_card, f"nome PT palavras '{pt_words_name}'")
            if matched:
                return matched

        namebar_en = self._fuzzy_namebar(ocr.namebar_text)
        namebar_line = best_namebar_line(ocr.namebar_text) or normalize_ocr_name(ocr.namebar_text)
        if namebar_en and (
            self._distinctive_name_in_raw_text(namebar_en, ocr.raw_text)
            or self._card_name_matches(namebar_en, namebar_line)
        ):
            matched = try_match(namebar_en, "nome superior OCR")
            if matched:
                return matched

        card, set_code, collector_number = self._find_by_collector_fraction(ocr.raw_text, ctx)
        matched = try_match(card, f"fração collector {set_code} #{collector_number}")
        if matched:
            return matched

        if not ctx.namebar_resolved:
            card, matched_name = self._fuzzy_portuguese_text(ocr.raw_text)
            matched = try_match(card, f"nome PT OCR '{matched_name}'")
            if matched:
                return matched

        card, matched_line = self._exact_english_self_reference_in_text(ocr.raw_text)
        matched = try_match(card, f"auto-referencia OCR '{matched_line}'")
        if matched:
            return matched

        card, matched_line = self._exact_english_title_in_text(ocr.raw_text)
        matched = try_match(card, f"título OCR bruto '{matched_line}'")
        if matched:
            return matched

        card, matched_line = self._fuzzy_english_title_in_text(ocr.raw_text)
        matched = try_match(card, f"título OCR aproximado '{matched_line}'")
        if matched:
            return matched

        card, matched_line = self._exact_english_name_in_text(ocr.raw_text)
        matched = try_match(card, f"nome inglês OCR '{matched_line}'")
        if matched:
            return matched

        card, matched_line = self._find_by_text_signature(ocr.raw_text)
        if card:
            return finalize(card, f"assinatura texto+artista '{matched_line}'")

        if ocr.name_hint and self._likely_name_hint(ocr.name_hint):
            normalized_hint = normalize_ocr_name(ocr.name_hint)
            card = None
            if is_plausible_namebar(normalized_hint):
                card = self._fuzzy_name(normalized_hint, strict_words=False, min_score=76)
            if not card and is_plausible_namebar(ocr.name_hint):
                card = self._fuzzy_name(ocr.name_hint, strict_words=True, min_score=84)
            if card and not self._hint_supports_card(card, ocr.name_hint):
                card = None
            matched = try_match(card, f"nome OCR '{ocr.name_hint}'")
            if matched:
                return matched

        card, matched_line = self._fuzzy_ocr_text(ocr.raw_text)
        matched = try_match(card, f"texto OCR '{matched_line}'")
        if matched:
            return matched
        raise LookupError("Não consegui encontrar a carta na base local.")

    @staticmethod
    def _card_name_matches(card: dict, name_hint: str) -> bool:
        card_name = card.get("name", "")
        if not card_name or not name_hint:
            return False
        if normalize_text(card_name) in normalize_text(name_hint):
            return True
        if fuzz:
            return fuzz.WRatio(card_name, name_hint) >= 72
        import difflib

        return difflib.SequenceMatcher(None, normalize_text(card_name), normalize_text(name_hint)).ratio() >= 0.72

    def _print_from_context(
        self, same_oracle_cards: list[dict], raw_text: str, *, include_tolerant_year: bool = True
    ) -> dict | None:
        """Pick a printing from footer/copyright evidence other than an exact
        collector match: the set card-count total (e.g. ".../81" -> the only
        printing in an 81-card set) and then the copyright year (e.g. the
        "© 1993-2011" range disambiguates the 2011 printing)."""
        if not same_oracle_cards:
            return None
        paper_cards = [
            item
            for item in same_oracle_cards
            if not item.get("digital") and "paper" in (item.get("games") or [])
        ]
        context_cards = paper_cards or same_oracle_cards
        totals = {total for _collector, total in self._collector_fractions_in_text(raw_text)}
        for total in totals:
            count_sets = set(self._sets_for_card_count(total))
            matches = [item for item in context_cards if (item.get("set") or "").upper() in count_sets]
            if len(matches) == 1:
                return matches[0]
        years = re.findall(r"\b(19\d{2}|20\d{2})\b", raw_text)
        release_years = {(item.get("released_at") or "")[:4] for item in context_cards}
        exact_years = [year for year in years if year in release_years]
        if exact_years:
            counts = {year: exact_years.count(year) for year in set(exact_years)}
            target = max(counts, key=lambda year: (counts[year], year))
            matches = [item for item in context_cards if (item.get("released_at") or "")[:4] == target]
            if len(matches) == 1:
                return matches[0]
        # The copyright year is often the only thing separating reprints, and it
        # sits on the lowest-contrast line, so OCR mangles a digit (e.g. "2010"
        # read as "1010"). Recover it by matching each printing's release year
        # against the OCR allowing one wrong digit.
        if include_tolerant_year:
            release_years = {year for year in release_years if re.fullmatch(r"\d{4}", year)}
            tolerant = self._years_in_copyright(raw_text, release_years)
            if len(tolerant) == 1:
                target = next(iter(tolerant))
                matches = [item for item in context_cards if (item.get("released_at") or "")[:4] == target]
                if len(matches) == 1:
                    return matches[0]
        return None

    def _flavor_scores(self, cards: list[dict], raw_text: str) -> dict[int, tuple[int, int]]:
        """(proper-noun hits, content-word hits) of each card's flavor text in
        the OCR. Flavor is stored in English, but proper nouns in an attribution
        ("—Sheoldred, Whispering One") are spelled the same across languages, so
        they still anchor a match against a Portuguese scan."""
        raw_words = set(normalize_text(raw_text).split())
        scores: dict[int, tuple[int, int]] = {}
        for item in cards:
            flavor = card_faces_text(item, "flavor_text")
            if not flavor:
                scores[id(item)] = (0, 0)
                continue
            proper = {
                normalize_text(word) for word in re.findall(r"[A-Z][A-Za-z'-]{3,}", flavor)
            }
            proper = {word for word in proper if word}
            proper_hits = 0
            for token in proper:
                if token in raw_words:
                    proper_hits += 1
                elif fuzz and any(fuzz.ratio(token, word) >= 85 for word in raw_words):
                    proper_hits += 1
            content_words = {word for word in normalize_text(flavor).split() if len(word) >= 4}
            content_hits = sum(1 for word in content_words if word in raw_words)
            scores[id(item)] = (proper_hits, content_hits)
        return scores

    def _flavor_compatible_prints(self, cards: list[dict], raw_text: str) -> list[dict]:
        """Drop printings whose flavor text contradicts what the OCR read. Many
        reprints differ only by flavor (Trained Armodon's Tempest quote is "These
        are its last days...", but the reprints say "Armodons are trained to step
        on things"), so a printing whose flavor is absent from the scan cannot be
        the card in hand. Only filters when at least one printing's flavor is
        strongly present, so a flavorless or badly garbled scan is left intact."""
        if len(cards) <= 1:
            return cards
        scores = self._flavor_scores(cards, raw_text)
        best = max(scores.values())
        if best[0] < 1 and best[1] < 3:
            return cards
        keep = [item for item in cards if scores[id(item)] == best]
        return keep or cards

    @staticmethod
    def _artist_compatible_prints(cards: list[dict], raw_text: str) -> list[dict]:
        if len(cards) <= 1:
            return cards
        raw = normalize_text(raw_text)
        raw_words = [word for word in raw.split() if len(word) >= 3]
        scores: dict[int, int] = {}
        for item in cards:
            artist = normalize_text(item.get("artist", ""))
            if not artist:
                scores[id(item)] = 0
                continue
            if artist in raw:
                scores[id(item)] = 2
                continue
            artist_words = [word for word in artist.split() if len(word) >= 3]
            if artist_words and fuzzy_word_count(artist_words, raw_words, threshold=88) == len(artist_words):
                scores[id(item)] = 1
            else:
                scores[id(item)] = 0
        best = max(scores.values(), default=0)
        if best <= 0:
            return cards
        keep = [item for item in cards if scores[id(item)] == best]
        return keep or cards

    @staticmethod
    def _years_in_copyright(raw_text: str, candidate_years: set[str]) -> set[str]:
        normalized = re.sub(r"[OoQ]", "0", raw_text)
        normalized = re.sub(r"[lI|]", "1", normalized)
        tokens = set(re.findall(r"(?<!\d)\d{4}(?!\d)", normalized))
        found = set()
        for year in candidate_years:
            for token in tokens:
                if year[:2] != token[:2]:
                    continue
                if sum(a != b for a, b in zip(year, token)) <= 1:
                    found.add(year)
                    break
        return found

    @staticmethod
    def _earliest_print(same_oracle_cards: list[dict], card: dict) -> dict:
        if not same_oracle_cards:
            return card
        return sorted(same_oracle_cards, key=lambda item: item.get("released_at") or "9999-99-99")[0]

    def _prefer_print_from_ocr(self, card: dict, raw_text: str) -> dict:
        oracle_id = card.get("oracle_id")
        set_codes_in_text = self._set_codes_in_text(raw_text)
        same_oracle_cards = [
            item
            for item in self.cards
            if item.get("lang") == "en" and item.get("oracle_id") == oracle_id
        ]
        collectors = self._collector_numbers_in_text(raw_text)
        pt_card = self.portuguese_for(card)
        preferred_set = (
            (pt_card or {}).get("preferred_set")
            or PREFERRED_PRINT_SETS.get(card.get("name", ""), "")
        ).upper()
        meaningful_collectors = {
            collector
            for collector in collectors
            if (re.sub(r"\D", "", collector) and int(re.sub(r"\D", "", collector)) >= 20)
        }
        if preferred_set and not meaningful_collectors:
            preferred = [
                item
                for item in same_oracle_cards
                if (item.get("set") or "").upper() == preferred_set
            ]
            if preferred:
                return preferred[0]

        def set_in_text(item: dict) -> bool:
            code = (item.get("set") or "").upper()
            return bool(code) and code in set_codes_in_text

        set_matches = [item for item in same_oracle_cards if set_in_text(item)]
        if set_matches:
            # When a collector number was read from the card, keep the printing
            # whose set AND collector both appear in the OCR text. Otherwise an
            # exact set+collector match gets overwritten by an arbitrary same-set
            # sibling (a serialized or alternate-art reprint under a different
            # collector number), which is how MSC #58 turned into MSC #369.
            if collectors:
                for item in set_matches:
                    if collector_key(item.get("collector_number")) in collectors:
                        return item
                if card in set_matches:
                    return card
            return set_matches[0]

        strong_context_pick = self._print_from_context(
            same_oracle_cards, raw_text, include_tolerant_year=False
        )
        if strong_context_pick and not collectors:
            return strong_context_pick

        # With no readable set/collector/year, the flavor text on the card is
        # the strongest remaining signal for which reprint this is: discard the
        # printings whose flavor contradicts the scan before falling back.
        compatible = self._flavor_compatible_prints(same_oracle_cards, raw_text)
        compatible = self._artist_compatible_prints(compatible, raw_text)
        context_pick = self._print_from_context(compatible, raw_text)
        if not collectors:
            if context_pick:
                return context_pick
            return self._earliest_print(compatible, card)
        current_collector = collector_key(card.get("collector_number"))
        if current_collector in collectors:
            return card
        same_cards = [
            item
            for item in self.cards
            if item.get("lang") == "en"
            and item.get("oracle_id") == oracle_id
            and collector_key(item.get("collector_number")) in collectors
        ]
        if same_cards:
            for item in same_cards:
                if (item.get("set") or "").upper() in set_codes_in_text:
                    return item
            return same_cards[0]
        return context_pick or self._earliest_print(compatible, card)

    @staticmethod
    def _collector_numbers_in_text(raw_text: str) -> set[str]:
        text = raw_text.upper()
        collectors = set()
        for match in re.finditer(r"\b([0-9OIL]{1,4}[A-Z]?)\s*/\s*[0-9OILSE]{1,5}\b", text):
            collector = ocr_collector_key(match.group(1))
            total_raw = match.group(0).split("/", 1)[1]
            total_raw = total_raw.translate(str.maketrans({"O": "0", "I": "1", "L": "1", "E": "8", "S": "5"}))
            total_match = re.match(r"\s*(\d{1,5})", total_raw)
            if total_match and ScryfallDatabase._is_likely_collector_fraction(collector, int(total_match.group(1))):
                collectors.add(collector)
        tokens = re.findall(r"[A-Z0-9]+", text)
        for index, token in enumerate(tokens[:-1]):
            normalized_token = token.translate(str.maketrans({"O": "0", "I": "1", "L": "1"}))
            normalized_next = tokens[index + 1].translate(str.maketrans({"O": "0", "I": "1", "L": "1", "E": "8", "S": "5"}))
            if re.fullmatch(r"0*\d{1,4}", normalized_token) and re.fullmatch(r"\d{1,5}", normalized_next):
                if not re.search(r"\d", tokens[index + 1]):
                    continue
                total = int(normalized_next)
                if total > 999:
                    continue
                number_match = re.match(r"0*(\d+)", normalized_token)
                collector = collector_key(normalized_token)
                if number_match and ScryfallDatabase._is_likely_collector_fraction(collector, total):
                    collectors.add(collector)
            right_window = tokens[index + 1 : index + 6]
            if (
                re.fullmatch(r"S\d{1,3}[A-Z]?", token)
                and any(value in {"C", "U", "R", "M"} for value in right_window[:2])
                and any(value in {"EN", "PT", "ES", "FR", "DE", "IT", "JP", "KO"} for value in right_window)
            ):
                collectors.add(collector_key(f"3{token[1:]}"))
        return collectors

    def _find_print_in_ocr_text(self, raw_text: str) -> tuple[dict | None, str, str]:
        if not raw_text:
            return None, "", ""
        known_sets = self._known_set_codes()
        fractions = self._collector_fractions_in_text(raw_text)
        fraction_totals = {collector: total for collector, total in fractions}
        language_codes = {"EN", "PT", "ES", "FR", "DE", "IT", "JP", "KO", "RU", "ZHS", "ZHT"}
        rarity_codes = {"C", "U", "R", "M"}
        tokens = re.findall(r"[A-Z0-9]+", raw_text.upper())

        def candidate_from(set_code: str, collector: str, allow_context_match: bool = False) -> dict | None:
            candidates = self._print_candidates(set_code, collector)
            if not candidates:
                return None
            for card in candidates:
                if self._card_name_in_raw_text(card, raw_text):
                    return card
            total = fraction_totals.get(collector, 0)
            collector_digits = re.sub(r"\D", "", collector)
            if collector_digits and int(collector_digits) <= 9 and not total:
                return None
            if total:
                if set_code in self._sets_for_card_count(total):
                    return self._pick_best_print_candidate(candidates, raw_text)
                return None
            if not allow_context_match:
                return None
            if self.by_print.get((set_code, collector)):
                return self._pick_best_print_candidate(candidates, raw_text)
            return None

        for index, token in enumerate(tokens):
            if token not in rarity_codes:
                continue
            collector_index = index + 1
            if collector_index < len(tokens) and tokens[collector_index] in rarity_codes:
                collector_index += 1
            if collector_index >= len(tokens):
                continue
            for collector in self._collector_token_variants(tokens[collector_index]):
                total = fraction_totals.get(collector, 0)
                for lookahead in range(collector_index + 1, min(collector_index + 25, len(tokens) - 1)):
                    set_code = self._corrected_set(tokens[lookahead], known_sets, collector, total)
                    if not set_code:
                        continue
                    if tokens[lookahead + 1] not in language_codes and (set_code, collector) not in self.by_print:
                        continue
                    if self._is_guessed_set_code(set_code, tokens[lookahead]) and not (
                        bool(total) or self._footer_shape_context(tokens, lookahead)
                    ):
                        continue
                    card = candidate_from(set_code, collector, True)
                    if card:
                        return card, set_code, collector
                    for corrected_collector in self._small_set_collector_variants(set_code, collector):
                        card = candidate_from(set_code, corrected_collector, True)
                        if card:
                            return card, set_code, corrected_collector

        checked = set()
        collector_indexes = list(range(len(tokens)))
        collector_indexes.sort(
            key=lambda item: 0 if re.fullmatch(r"S\d{1,3}[A-Z]?", tokens[item]) else 1
        )
        for index in collector_indexes:
            token = tokens[index]
            collector_variants = self._collector_token_variants(token)
            if not collector_variants:
                continue
            if token.isdigit() and len(token) <= 2:
                previous_is_number = index > 0 and tokens[index - 1].isdigit()
                next_is_number = index + 1 < len(tokens) and tokens[index + 1].isdigit()
                if previous_is_number or next_is_number:
                    continue
            if index + 1 < len(tokens) and tokens[index + 1].isdigit():
                search_start = index + 2
            else:
                search_start = index + 1
            for collector in collector_variants:
                total = fraction_totals.get(collector, 0)
                for lookahead in range(search_start, min(search_start + 25, len(tokens) - 1)):
                    set_code = self._corrected_set(tokens[lookahead], known_sets, collector, total)
                    if not set_code:
                        continue
                    key = (set_code, collector)
                    if key in checked:
                        continue
                    checked.add(key)
                    collector_digits = re.sub(r"\D", "", collector)
                    fixed_print_context = bool(
                        self.OCR_SET_FIXES.get(tokens[lookahead])
                        and self.by_print.get(key)
                        and collector_digits
                        and int(collector_digits) >= 20
                    )
                    has_print_context = (
                        bool(total)
                        or self._set_token_has_print_context(tokens, lookahead)
                        or fixed_print_context
                    )
                    # A fuzzy-guessed code (rules-text "once" -> ONE, "Cont"
                    # -> CON) needs true footer shape around the token; the
                    # raw token's own shape proves nothing about a guess.
                    if (
                        has_print_context
                        and not bool(total)
                        and not fixed_print_context
                        and self._is_guessed_set_code(set_code, tokens[lookahead])
                        and not self._footer_shape_context(tokens, lookahead)
                    ):
                        has_print_context = False
                    card = candidate_from(set_code, collector, has_print_context)
                    if card:
                        return card, set_code, collector
                    for corrected_collector in self._small_set_collector_variants(set_code, collector):
                        corrected_key = (set_code, corrected_collector)
                        if corrected_key in checked:
                            continue
                        checked.add(corrected_key)
                        card = candidate_from(set_code, corrected_collector, has_print_context)
                        if card:
                            return card, set_code, corrected_collector
        return None, "", ""

    def _card_name_in_raw_text(self, card: dict, raw_text: str) -> bool:
        names = [card.get("name", "")]
        for face in card.get("card_faces") or []:
            face_name = face.get("name") or ""
            if face_name:
                names.append(face_name)
        pt_card = self.portuguese_for(card)
        if pt_card:
            names.append(pt_card.get("printed_name") or pt_card.get("name") or "")
            printed_name = (pt_card or {}).get("printed_name") or ""
            for part in re.split(r"\s*/\s*", printed_name):
                part = part.strip()
                if part:
                    names.append(part)
        names.extend(self._foreign_names_for(card))
        names = [name for name in names if name]
        if not names:
            return False
        normalized_raw = normalize_text(raw_text)
        for card_name in names:
            normalized_name = normalize_text(card_name)
            if normalized_phrase_in_text(normalized_name, normalized_raw):
                return True
            name_words = [word for word in normalized_name.split() if len(word) >= 4]
            required_matches = 2 if len(name_words) >= 2 else 1
            raw_words = [word for word in normalized_raw.split() if len(word) >= 3]
            if name_words:
                if fuzzy_word_count(name_words, raw_words) >= required_matches:
                    return True
                joined_raw = " ".join(raw_words)
                joined_name = " ".join(name_words)
                if not fuzz or fuzz.partial_ratio(joined_name, joined_raw) < 78:
                    continue
            if fuzz:
                for line in raw_text.splitlines():
                    cleaned = clean_ocr_line(line)
                    normalized_line = normalize_text(cleaned)
                    line_words = normalized_line.split()
                    if len(line_words) > 8:
                        continue
                    if name_words and fuzzy_word_count(name_words, [word for word in line_words if len(word) >= 3]) < required_matches:
                        continue
                    if fuzz.WRatio(card_name, cleaned) >= 84:
                        return True
        return False

    def _strong_card_name_in_raw_text(self, card: dict, raw_text: str) -> bool:
        names = [card.get("name", "")]
        for face in card.get("card_faces") or []:
            face_name = face.get("name") or ""
            if face_name:
                names.append(face_name)
        pt_card = self.portuguese_for(card)
        if pt_card:
            printed_name = pt_card.get("printed_name") or pt_card.get("name") or ""
            names.append(printed_name)
            names.extend(part.strip() for part in re.split(r"\s*/\s*", printed_name) if part.strip())
        names.extend(self._foreign_names_for(card))

        normalized_raw = normalize_text(raw_text)
        raw_words = [word for word in normalized_raw.split() if len(word) >= 3]
        for card_name in (name for name in names if name):
            normalized_name = normalize_text(card_name)
            if normalized_phrase_in_text(normalized_name, normalized_raw):
                return True
            name_words = [
                word
                for word in normalized_name.split()
                if len(word) >= 3 and word not in COMMON_EN_OCR_WORDS and word not in COMMON_PT_OCR_WORDS
            ]
            if len(name_words) < 2:
                continue
            required = min(3, len(name_words))
            if fuzzy_word_count(name_words, raw_words, threshold=88) >= required:
                return True
        return False

    def _hint_card_shares_set(self, hint_card: dict, set_code: str) -> bool:
        if not set_code or not hint_card:
            return False
        oracle_id = hint_card.get("oracle_id") or hint_card.get("name")
        prints = self.en_by_oracle.get(oracle_id) or self.pt_by_oracle.get(oracle_id) or []
        return any((item.get("set") or "").upper() == set_code for item in prints)

    def is_confident_match(self, card: dict, ctx: OcrMatchContext) -> bool:
        ocr = ctx.ocr
        name_present = self._card_name_in_raw_text(card, ocr.raw_text)
        is_hint_card = bool(ctx.hint_card and ctx.hint_card.get("oracle_id") == card.get("oracle_id"))
        exact_title_card, _ = self._exact_english_title_in_text(ocr.namebar_text)
        has_exact_title = bool(
            exact_title_card and exact_title_card.get("oracle_id") == card.get("oracle_id")
        )
        exact_raw_title_card, _ = self._exact_english_title_in_text(ocr.raw_text)
        has_exact_raw_title = bool(
            exact_raw_title_card and exact_raw_title_card.get("oracle_id") == card.get("oracle_id")
            and self._best_oracle_support(card, ocr.raw_text) >= 3
        )
        self_reference_card, _ = self._exact_english_self_reference_in_text(ocr.raw_text)
        has_self_reference = bool(
            self_reference_card and self_reference_card.get("oracle_id") == card.get("oracle_id")
        )
        fuzzy_title_card, _ = self._fuzzy_english_title_in_text(ocr.raw_text)
        fuzzy_namebar_title_card, _ = self._fuzzy_english_title_in_text(ocr.namebar_text)
        has_fuzzy_title = bool(
            (
                fuzzy_title_card
                and fuzzy_title_card.get("oracle_id") == card.get("oracle_id")
            )
            or (
                fuzzy_namebar_title_card
                and fuzzy_namebar_title_card.get("oracle_id") == card.get("oracle_id")
            )
        )
        exact_name_card, _ = self._exact_english_name_in_text(ocr.raw_text)
        has_exact_name = bool(
            exact_name_card and exact_name_card.get("oracle_id") == card.get("oracle_id")
        )
        # The foreign-namebar matcher applies its own strict gates (full and
        # compact ratio plus an anchor word), so agreeing with it is strong
        # evidence — often the only kind an old foreign card can produce.
        foreign_namebar_card, _ = self._fuzzy_foreign_namebar(ocr.namebar_text)
        has_foreign_namebar = bool(
            foreign_namebar_card and foreign_namebar_card.get("oracle_id") == card.get("oracle_id")
        )
        # When this card's own name is the title read clearly from the image, a
        # collector number that carries an OCR digit error (e.g. 215 read as 219)
        # must not veto it — the name is the stronger signal.
        strong_name = (
            (name_present and is_hint_card)
            or has_exact_title
            or has_exact_raw_title
            or has_self_reference
            or has_fuzzy_title
            or has_exact_name
            or has_foreign_namebar
        )
        strong_print = self._print_evidence_in_ocr(card, ctx) or self._print_pair_occurrences(card, ctx) >= 2
        if not strong_name and self._ocr_contradicts_card(card, ctx):
            return False
        if (
            ctx.hint_card
            and ctx.hint_card.get("oracle_id") != card.get("oracle_id")
            and not strong_name
        ):
            if ctx.hint_from_title:
                return False
            if not strong_print:
                return False
            if not self._card_name_in_raw_text(card, ocr.raw_text):
                # A name read clearly from the card must outweigh a collector
                # number when the named card also exists in this set: that is the
                # signature of a misread digit (e.g. 215 -> 219 turning
                # "Zhur-Taa Goblin" into "Senate Griffin").
                if self._hint_card_shares_set(ctx.hint_card, (card.get("set") or "").upper()):
                    return False
        if strong_print:
            return True
        if has_exact_title or has_exact_raw_title or has_self_reference or has_exact_name:
            return True
        if has_foreign_namebar:
            return True
        if has_fuzzy_title and (
            self._distinctive_name_in_raw_text(card, ocr.raw_text)
            or self._artist_name_in_raw_text(
                card,
                normalize_text(ocr.raw_text),
                re.sub(r"\s+", "", normalize_text(ocr.raw_text)),
                [word for word in normalize_text(ocr.raw_text).split() if len(word) >= 3],
            )
        ):
            return True
        if self._print_evidence_in_ocr(card, ctx):
            return True
        if self._card_name_in_raw_text(card, ocr.raw_text):
            return True
        if ctx.hint_card and ctx.hint_card.get("oracle_id") == card.get("oracle_id"):
            return True
        set_code = (card.get("set") or "").upper()
        collector = collector_key(card.get("collector_number"))
        for index, token in enumerate(ctx.tokens):
            if collector not in self._collector_token_variants(token):
                continue
            window = set(ctx.tokens[index + 1 : index + 16])
            if set_code in window:
                return True
        for fraction_collector, total in ctx.fractions:
            if fraction_collector != collector:
                continue
            matching_sets = self._sets_for_card_count(total)
            if set_code not in matching_sets:
                continue
            if set_code in ctx.sets_in_text:
                return True
            if len(matching_sets) == 1 and set_code in ctx.tokens:
                return True
        card_type = normalize_text(card_faces_text(card, "type_line"))
        card_text = normalize_text(card_faces_text(card, "oracle_text"))
        raw = normalize_text(ocr.raw_text)
        type_words = [word for word in card_type.split() if len(word) >= 5]
        text_words = [word for word in card_text.split() if len(word) >= 7]
        return (
            sum(1 for word in type_words if word in raw) >= 2
            and sum(1 for word in text_words if word in raw) >= 4
        )

    def _ocr_contradicts_card(self, card: dict, ctx: OcrMatchContext) -> bool:
        raw_text = ctx.ocr.raw_text
        fractions = ctx.fractions
        collectors = self._collector_numbers_in_text(raw_text)
        if not fractions and not collectors:
            return False
        card_collector = collector_key(card.get("collector_number"))
        card_set = (card.get("set") or "").upper()
        oracle_id = card.get("oracle_id")

        fraction_collectors = {collector for collector, _total in fractions}
        if card_collector in fraction_collectors:
            if card_set in ctx.sets_in_text:
                return False
            if self._card_name_in_raw_text(card, raw_text):
                return False
            supporting = False
            for collector, total in fractions:
                if collector != card_collector:
                    continue
                if card_set in self._sets_for_card_count(total):
                    supporting = True
                    break
            return not supporting

        if card_collector in collectors and self._oracle_has_collector(oracle_id, card_collector):
            return False

        if fraction_collectors:
            for collector, total in fractions:
                if self._oracle_has_collector(oracle_id, collector):
                    continue
                matching_sets = self._sets_for_card_count(total)
                for set_code in matching_sets:
                    if self.by_print.get((set_code, collector)):
                        return True

        if collectors and not self._oracle_has_collector(oracle_id, card_collector):
            for collector in collectors:
                if self._oracle_has_collector(oracle_id, collector):
                    continue
                if collector != card_collector:
                    return True
        return False

    @staticmethod
    def _likely_name_hint(name_hint: str) -> bool:
        normalized = normalize_text(name_hint)
        words = normalized.split()
        if len(words) < 1 or len(words) > 6:
            return False
        if {
            "hlus",
            "iilug",
            "illus",
            "iilus",
            "ilus",
            "inus",
            "tlug",
            "illustrated",
            "illustration",
            "hust",
            "tilug",
            "tilus",
            "titus",
            "tllus",
            "tlus",
            "tlust",
            "mlus",
            "mluse",
            "wizards",
            "coast",
            "rights",
            "reserved",
        }.intersection(words):
            return False
        rules_words = {
            "changeling",
            "this",
            "card",
            "every",
            "creature",
            "creatures",
            "type",
            "whenever",
            "target",
            "until",
            "each",
            "player",
            "players",
            "play",
            "bury",
            "cannot",
            "spend",
            "turn",
            "number",
            "snow",
            "covered",
            "swamps",
            "magic",
            "gathering",
            "colecao",
            "raridade",
            "descricao",
            "sacrifique",
            "sacrifice",
            "causa",
            "dano",
            "damage",
            "creatures",
            "bury",
            "cannot",
            "spend",
            "ataca",
            "ataque",
            "bloquear",
        }
        hard_action_words = {
            "sacrifique",
            "sacrifice",
            "causa",
            "dano",
            "damage",
            "ataca",
            "ataque",
            "bloquear",
            "serving",
            "nothing",
            "posters",
            "navy",
        }
        if hard_action_words.intersection(words):
            return False
        return len(rules_words.intersection(words)) <= 1

    def _fuzzy_name(
        self,
        name_hint: str,
        strict_words: bool = False,
        min_score: int = 84,
        prefer_en: bool = False,
    ) -> dict | None:
        normalized_hint = normalize_text(name_hint)
        if not normalized_hint:
            return None
        if process and fuzz:
            pt_card = None
            pt_score = 0
            pt_name = ""
            en_card = None
            en_score = 0
            en_name = ""
            pt_choices = self._pt_norm_choices()
            if pt_choices:
                pt_match = process.extractOne(normalized_hint, pt_choices.keys(), scorer=fuzz.WRatio)
                if pt_match and pt_match[1] >= min_score:
                    pt_name, pt_index = pt_choices[pt_match[0]]
                    pt_card = self.cards[pt_index]
                    pt_score = pt_match[1]
            choices = self._en_norm_choices()
            en_match = process.extractOne(normalized_hint, choices.keys(), scorer=fuzz.WRatio) if choices else None
            if en_match and en_match[1] >= min_score:
                candidate_name, candidate_index = choices[en_match[0]]
                if not strict_words or self._name_words_present(candidate_name, normalized_hint):
                    en_card = self.cards[candidate_index]
                    en_score = en_match[1]
                    en_name = candidate_name
            # A name that only explains part of what was read (e.g. the single-word
            # plane "Naya") must not win over a candidate that explains more of it
            # (e.g. "Medalhão de Naya"), even when English is normally preferred.
            if en_card and pt_card and en_card.get("oracle_id") != pt_card.get("oracle_id"):
                en_cov = self._hint_word_coverage(en_name, normalized_hint)
                pt_cov = self._hint_word_coverage(pt_name, normalized_hint)
                if pt_cov > en_cov:
                    return pt_card
                if en_cov > pt_cov:
                    return en_card
            if en_card and (not pt_card or en_score >= pt_score + 3 or prefer_en):
                return en_card
            if pt_card and (not en_card or pt_score > en_score + 3):
                return pt_card
            return en_card or pt_card

        import difflib

        names = [name for name, _index in self.name_choices]
        match = difflib.get_close_matches(name_hint, names, n=1, cutoff=0.72)
        if match:
            if strict_words and not self._name_words_present(match[0], normalized_hint):
                return None
            index = next(index for name, index in self.name_choices if name == match[0])
            return self.cards[index]
        pt_names = [name for name, _index in self.pt_name_choices]
        pt_match = difflib.get_close_matches(name_hint, pt_names, n=1, cutoff=0.72)
        if pt_match:
            index = next(index for name, index in self.pt_name_choices if name == pt_match[0])
            return self.cards[index]
        return None

    @staticmethod
    def _name_words_present(card_name: str, normalized_text: str) -> bool:
        words = [word for word in normalize_text(card_name).split() if len(word) >= 4]
        if not words:
            return False
        required = 2 if len(words) >= 2 else 1
        return sum(1 for word in words if word in normalized_text) >= required

    @staticmethod
    def _hint_word_coverage(card_name: str, normalized_hint: str) -> int:
        """How many significant words of ``card_name`` appear in the read text.

        Used to decide, between two candidate names, which one better explains
        what the OCR actually read — so a short substring match cannot beat a
        fuller one.
        """
        name_words = [word for word in normalize_text(card_name).split() if len(word) >= 3]
        hint_words = [word for word in (normalized_hint or "").split() if len(word) >= 3]
        if not name_words or not hint_words:
            return 0
        return fuzzy_word_count(name_words, hint_words, threshold=85)

    def _fuzzy_portuguese_text(
        self,
        raw_text: str,
        *,
        namebar_mode: bool = False,
        max_lines: int = 20,
    ) -> tuple[dict | None, str]:
        if not raw_text or not process or not fuzz or not self.pt_name_choices:
            return None, ""
        if not namebar_mode:
            short_lines = []
            for line in raw_text.splitlines():
                cleaned = clean_ocr_line(line)
                words = normalize_text(cleaned).split()
                if not words or len(words) > 8:
                    continue
                short_lines.append(line)
                if len(short_lines) >= max_lines:
                    break
            if not short_lines:
                return None, ""
            raw_text = "\n".join(short_lines)
        raw_words = [word for word in normalize_text(raw_text).split() if len(word) >= 3]
        if namebar_mode:
            compact_raw = re.sub(r"\s+", "", normalize_text(raw_text))
            # Lines whose only significant token is a single word: the natural
            # shape of a one-word card name ("Gelar", "Aríete") in the title
            # crop. Compared token-to-token so a short name is never found
            # inside a longer word ("gelar" inside "congelar").
            single_line_words = []
            for namebar_line in raw_text.splitlines():
                line_words = [
                    word
                    for word in normalize_text(clean_ocr_line(namebar_line)).split()
                    if len(word) >= 3
                ]
                if len(line_words) == 1 and len(line_words[0]) >= 4:
                    single_line_words.append(line_words[0])
            best_word_card = None
            best_word_name = ""
            best_word_score = 0
            for pt_name, index in self.pt_name_choices:
                normalized_pt_name = normalize_text(pt_name)
                pt_words = [word for word in normalized_pt_name.split() if len(word) >= 4]
                if not pt_words:
                    continue
                compact_pt_name = re.sub(r"\s+", "", normalized_pt_name)
                if len(compact_pt_name) >= 8 and compact_raw:
                    compact_score = 0
                    if compact_pt_name in compact_raw:
                        compact_score = len(pt_words) * 120 + len(compact_pt_name)
                    elif fuzz:
                        compact_ratio = fuzz.partial_ratio(compact_pt_name, compact_raw)
                        if compact_ratio >= 94:
                            compact_score = len(pt_words) * 110 + len(compact_pt_name)
                        elif (
                            compact_ratio >= 80
                            and len(pt_words) == 1
                            and len(raw_words) <= 3
                        ):
                            compact_score = int(compact_ratio) + len(compact_pt_name)
                    if compact_score > best_word_score:
                        best_word_score = compact_score
                        best_word_name = pt_name
                        best_word_card = self.cards[index]
                        continue
                elif (
                    single_line_words
                    and len(pt_words) == 1
                    and 5 <= len(compact_pt_name) <= 7
                    and compact_pt_name not in COMMON_PT_OCR_WORDS
                    and compact_pt_name not in MTG_KEYWORD_NAME_STOPWORDS
                ):
                    # Short one-word names never reach the compact branch
                    # (>= 8 chars); accept them only when a namebar line reads
                    # as exactly that word.
                    best_ratio = max(
                        (fuzz.ratio(compact_pt_name, word) for word in single_line_words),
                        default=0,
                    )
                    min_ratio = 82 if len(compact_pt_name) >= 6 else 88
                    if best_ratio >= min_ratio:
                        compact_score = int(best_ratio) + len(compact_pt_name)
                        if compact_score > best_word_score:
                            best_word_score = compact_score
                            best_word_name = pt_name
                            best_word_card = self.cards[index]
                            continue
                all_pt_words = [word for word in normalize_text(pt_name).split() if len(word) >= 3]
                first_word = all_pt_words[0] if all_pt_words else ""
                normalized_pt_tokens = normalize_text(pt_name).split()
                has_short_article = bool(
                    normalized_pt_tokens
                    and normalized_pt_tokens[0] in {"a", "o", "as", "os"}
                    and len(pt_words) == 1
                )
                if len(all_pt_words) > 1 and first_word:
                    first_raw_word = raw_words[0] if raw_words else ""
                    if not words_fuzzy_match(first_word, first_raw_word, 86):
                        if not (len(first_word) <= 2 and fuzzy_word_count(pt_words, raw_words, threshold=86)):
                            continue
                if len(pt_words) == 1 and len(raw_words) > 1:
                    if not (
                        (
                            (len(all_pt_words) >= 2 and len(first_word) <= 2)
                            or has_short_article
                        )
                        and fuzzy_word_count(pt_words, raw_words, threshold=86)
                    ):
                        continue
                matches = fuzzy_word_count(pt_words, raw_words)
                required = min(3, len(pt_words)) if len(pt_words) >= 3 else min(2, len(pt_words))
                if matches < required:
                    continue
                score = matches * 100 + min(len(pt_words), 6)
                if score > best_word_score:
                    best_word_score = score
                    best_word_name = pt_name
                    best_word_card = self.cards[index]
            if best_word_card:
                return best_word_card, best_word_name

        choices = {name: index for name, index in self.pt_name_choices}
        best_card = None
        best_name = ""
        best_score = 0
        for line in raw_text.splitlines():
            cleaned = clean_ocr_line(line)
            normalized = normalize_text(cleaned)
            words = [word for word in normalized.split() if len(word) >= 3]
            if not words or len(words) > 8:
                continue
            for size in range(min(5, len(words)), 1, -1):
                for start in range(0, len(words) - size + 1):
                    candidate = " ".join(words[start : start + size])
                    match = process.extractOne(candidate, choices.keys(), scorer=fuzz.WRatio)
                    if not match:
                        continue
                    pt_name = match[0]
                    pt_words = [word for word in normalize_text(pt_name).split() if len(word) >= 4]
                    all_pt_words = [word for word in normalize_text(pt_name).split() if len(word) >= 3]
                    first_word = all_pt_words[0] if all_pt_words else ""
                    if len(all_pt_words) > 1 and first_word:
                        if not any(words_fuzzy_match(first_word, word, 86) for word in words):
                            continue
                    if len(pt_words) == 1 and len(words) > 2:
                        continue
                    required = min(3, len(pt_words)) if len(pt_words) >= 3 else min(2, len(pt_words))
                    if pt_words and fuzzy_word_count(pt_words, words) < required:
                        continue
                    adjusted_score = match[1] + (len(pt_words) * 3)
                    if adjusted_score > best_score:
                        best_score = adjusted_score
                        best_name = pt_name
                        best_card = self.cards[choices[pt_name]]
        if best_card and best_score >= 90:
            return best_card, best_name
        return None, ""

    @staticmethod
    def _title_extra_words_are_noise(words: list[str], matched_size: int) -> bool:
        extras = words[matched_size:]
        if all(re.fullmatch(r"\d+|[0-9wubrgcxyz]+", word) for word in extras):
            return True
        if (
            len(extras) == 1
            and re.fullmatch(r"[a-z]{2,4}", extras[0])
            and extras[0] not in COMMON_EN_OCR_WORDS
            and extras[0] not in COMMON_PT_OCR_WORDS
        ):
            return True
        return False

    def _exact_english_title_in_text(self, raw_text: str) -> tuple[dict | None, str]:
        if not raw_text:
            return None, ""
        choices = self._en_norm_choices()
        lines_seen = 0
        ignored = COMMON_EN_OCR_WORDS | COMMON_PT_OCR_WORDS | {"again", "letter"}
        for line in raw_text.splitlines():
            cleaned = clean_ocr_line(line)
            if ":" in cleaned or re.search(r"[\u2013\u2014]", cleaned):
                continue
            words = title_words(cleaned)
            if not words or len(words) > 4:
                continue
            lines_seen += 1
            if lines_seen > 260:
                break
            for size in range(min(5, len(words)), 0, -1):
                candidate = " ".join(words[:size])
                if candidate not in choices:
                    continue
                first_word = words[0]
                # Multi-word exact matches may start with a short word ("Rod of
                # Ruin", "The Brute"); single-word candidates stay at >= 4.
                min_first = 3 if size >= 2 else 4
                if len(first_word) < min_first or first_word in ignored:
                    continue
                if size < len(words) and not self._title_extra_words_are_noise(words, size):
                    continue
                original_name, card_index = choices[candidate]
                if len(normalize_text(original_name).split()) == size:
                    return self.cards[card_index], cleaned
        return None, ""

    def _fuzzy_english_title_in_text(self, raw_text: str) -> tuple[dict | None, str]:
        if not raw_text:
            return None, ""
        import difflib

        choices = self._en_norm_choices()
        choice_keys = list(choices.keys())

        def extract_one(candidate: str, keys, scorer=None):
            if process and fuzz:
                return process.extractOne(candidate, keys, scorer=scorer or fuzz.ratio)
            best_key = ""
            best_score = 0.0
            for key in keys:
                score = difflib.SequenceMatcher(None, candidate, key).ratio() * 100
                if score > best_score:
                    best_key = key
                    best_score = score
            return (best_key, best_score, None) if best_key else None

        one_word_choices = {
            normalized_name: choice
            for normalized_name, choice in choices.items()
            if len(normalized_name.split()) == 1 and len(normalized_name) >= 5
        }
        compact_choices = {}
        for normalized_name, choice in choices.items():
            name_words = normalized_name.split()
            if len(name_words) < 2:
                continue
            compact_name = "".join(name_words)
            if len(compact_name) >= 8:
                compact_choices.setdefault(compact_name, choice)
        ignored = COMMON_EN_OCR_WORDS | COMMON_PT_OCR_WORDS | {
            "again",
            "illus",
            "letter",
            "wizards",
            "coast",
        }
        best_card = None
        best_line = ""
        best_score = 0.0
        lines_seen = 0
        normalized_raw = normalize_text(raw_text)
        compact_raw = re.sub(r"\s+", "", normalized_raw)
        raw_words = [word for word in normalized_raw.split() if len(word) >= 3]

        def best_supported_print(card: dict) -> tuple[dict, int]:
            oracle_id = card.get("oracle_id")
            same_oracle = [
                item
                for item in self.cards
                if item.get("lang") == "en" and item.get("oracle_id") == oracle_id
            ]
            candidates = same_oracle or [card]
            scored = [(self._rules_text_support_score(item, raw_text), item) for item in candidates]
            scored.sort(key=lambda item: item[0], reverse=True)
            support, supported_card = scored[0]
            return supported_card, support

        for line in raw_text.splitlines():
            cleaned = clean_ocr_line(line)
            if ":" in cleaned or re.search(r"[\u2013\u2014]", cleaned):
                continue
            # A digit surrounded by words is rules text ("deals 1 damage to
            # target"), never a title; mana-cost digits only trail the name.
            if re.search(r"[a-z]\s+\d+\s+[a-z]", normalize_text(cleaned)):
                continue
            words = [word for word in normalize_text(cleaned).split() if len(word) >= 3]
            if not words or len(words) > 4:
                continue
            lines_seen += 1
            if lines_seen > 260:
                break
            first_word = words[0]
            if len(first_word) < 4 or first_word in ignored:
                continue
            if len(words) == 1 and len(first_word) >= 5 and one_word_choices:
                match = extract_one(first_word, one_word_choices.keys(), scorer=fuzz.ratio if fuzz else None)
                if match and match[1] >= 88:
                    original_name, card_index = one_word_choices[match[0]]
                    card = self.cards[card_index]
                    card, support = best_supported_print(card)
                    if support >= 4:
                        score = float(match[1]) + support * 20 + 1
                        if score > best_score:
                            best_score = score
                            best_card = card
                            best_line = cleaned
            if len(first_word) >= 8 and compact_choices and process and fuzz:
                compact_match = extract_one(first_word, compact_choices.keys(), scorer=fuzz.ratio)
                if compact_match and compact_match[1] >= 76:
                    original_name, card_index = compact_choices[compact_match[0]]
                    card = self.cards[card_index]
                    card, support = best_supported_print(card)
                    if compact_match[1] < 94 and (
                        support < 6 or not self._card_name_in_raw_text(card, raw_text)
                    ):
                        continue
                    score = (
                        float(compact_match[1])
                        + len(normalize_text(original_name).split())
                        + 1
                        + support * 20
                    )
                    if score > best_score:
                        best_score = score
                        best_card = card
                        best_line = cleaned
            if len(words) >= 2:
                damaged_title_keys = [
                    normalized_name
                    for normalized_name in choice_keys
                    if len(normalized_name.split()) == len(words)
                    and len(normalized_name.split()[-1]) >= 5
                    and normalized_name.split()[-1] == words[-1]
                ]
                if damaged_title_keys:
                    candidate = " ".join(words)
                    damaged_match = extract_one(
                        candidate, damaged_title_keys, scorer=fuzz.ratio if fuzz else None
                    )
                    if damaged_match and damaged_match[1] >= 70:
                        _original_name, card_index = choices[damaged_match[0]]
                        card = self.cards[card_index]
                        card, support = best_supported_print(card)
                        artist_ok = self._artist_name_in_raw_text(
                            card, normalized_raw, compact_raw, raw_words
                        )
                        if artist_ok and support >= 5:
                            score = float(damaged_match[1]) + support * 20 + 20
                            if score > best_score:
                                best_score = score
                                best_card = card
                                best_line = cleaned
            for size in range(min(4, len(words)), 1, -1):
                candidate = " ".join(words[:size])
                candidate_words = candidate.split()
                candidate_keys = []
                for normalized_name in choice_keys:
                    name_words = normalized_name.split()
                    if len(name_words) > size:
                        continue
                    if len(name_words) < size and not self._title_extra_words_are_noise(candidate_words, len(name_words)):
                        continue
                    if not all(
                        candidate_word[:2] == name_word[:2]
                        or (
                            len(candidate_word) >= 4
                            and len(name_word) >= 4
                            and candidate_word[1:] == name_word[1:]
                        )
                        for candidate_word, name_word in zip(candidate_words[: len(name_words)], name_words)
                    ):
                        continue
                    candidate_keys.append(normalized_name)
                if not candidate_keys:
                    continue
                match = extract_one(candidate, candidate_keys, scorer=fuzz.ratio if fuzz else None)
                if not match or match[1] < 86:
                    continue
                original_name, card_index = choices[match[0]]
                name_words = normalize_text(original_name).split()
                if len(name_words) > size:
                    continue
                if len(name_words) < size and not self._title_extra_words_are_noise(candidate.split(), len(name_words)):
                    continue
                candidate_name_words = candidate.split()[: len(name_words)]
                if not all(
                    candidate_word[:2] == name_word[:2]
                    or (
                        len(candidate_word) >= 4
                        and len(name_word) >= 4
                        and candidate_word[1:] == name_word[1:]
                    )
                    for candidate_word, name_word in zip(candidate_name_words, name_words)
                ):
                    continue
                card = self.cards[card_index]
                card, support = best_supported_print(card)
                artist_ok = self._artist_name_in_raw_text(card, normalized_raw, compact_raw, raw_words)
                if match[1] < 88 and not artist_ok:
                    continue
                score = float(match[1]) + size + support * 20
                if artist_ok:
                    score += 25
                if score > best_score:
                    best_score = score
                    best_card = card
                    best_line = cleaned
        if best_card:
            return best_card, best_line
        return None, ""

    def _exact_english_name_in_text(self, raw_text: str) -> tuple[dict | None, str]:
        if not raw_text:
            return None, ""
        choices = self._en_norm_choices()
        best_card = None
        best_line = ""
        best_score = 0
        lines_seen = 0
        for line in raw_text.splitlines():
            cleaned = clean_ocr_line(line)
            words = normalize_text(cleaned).split()
            if not words or len(words) > 9:
                continue
            lines_seen += 1
            if lines_seen > 100:
                break
            for size in range(min(6, len(words)), 1, -1):
                for start in range(0, len(words) - size + 1):
                    candidate = " ".join(words[start : start + size])
                    if candidate not in choices:
                        continue
                    original_name, card_index = choices[candidate]
                    name_words = normalize_text(original_name).split()
                    if len(name_words) != size:
                        continue
                    distinctive_words = [
                        word
                        for word in name_words
                        if len(word) >= 4
                        and word not in COMMON_EN_OCR_WORDS
                        and word not in COMMON_PT_OCR_WORDS
                    ]
                    if not distinctive_words:
                        continue
                    card = self.cards[card_index]
                    support = self._rules_text_support_score(card, raw_text)
                    if support < 4:
                        continue
                    score = support * 100 + size
                    if score > best_score:
                        best_score = score
                        best_card = card
                        best_line = cleaned
        if best_card:
            return best_card, best_line
        return None, ""

    @staticmethod
    def _artist_name_in_raw_text(card: dict, normalized_raw: str, compact_raw: str, raw_words: list[str]) -> bool:
        artist = normalize_text(card.get("artist", ""))
        if not artist:
            return False
        if artist in normalized_raw:
            return True
        compact_artist = re.sub(r"\s+", "", artist)
        if len(compact_artist) >= 7 and compact_artist in compact_raw:
            return True
        artist_words = [word for word in artist.split() if len(word) >= 3]
        return bool(artist_words) and all(word in raw_words for word in artist_words)

    def _find_by_text_signature(self, raw_text: str) -> tuple[dict | None, str]:
        if not raw_text:
            return None, ""
        normalized_raw = normalize_text(raw_text)
        compact_raw = re.sub(r"\s+", "", normalized_raw)
        raw_words = [word for word in normalized_raw.split() if len(word) >= 3]
        best_by_oracle: dict[str, tuple[int, dict]] = {}
        for card in self.cards:
            if card.get("lang") != "en" or card.get("layout") in _NON_CATALOG_LAYOUTS:
                continue
            if not self._artist_name_in_raw_text(card, normalized_raw, compact_raw, raw_words):
                continue
            score = self._rules_text_support_score(card, raw_text)
            if score < 8:
                continue
            oracle_id = card.get("oracle_id") or card.get("name")
            current = best_by_oracle.get(oracle_id)
            if current is None or score > current[0]:
                best_by_oracle[oracle_id] = (score, card)
        if not best_by_oracle:
            return None, ""
        ranked = sorted(best_by_oracle.values(), key=lambda item: item[0], reverse=True)
        best_score, best_card = ranked[0]
        if best_score < 10:
            return None, ""
        if len(ranked) > 1 and ranked[1][0] > best_score - 3:
            return None, ""
        return best_card, best_card.get("artist", "")

    @staticmethod
    def _self_reference_line_context(words: list[str], index: int, name_size: int = 1) -> bool:
        following = words[index + name_size : index + name_size + 4]
        previous = words[max(0, index - 3) : index]
        action_words = {
            "attacks",
            "becomes",
            "blocks",
            "can",
            "cant",
            "cannot",
            "come",
            "comes",
            "costs",
            "deals",
            "dies",
            "enters",
            "exiles",
            "gains",
            "gets",
            "has",
            "leaves",
            "loses",
            "may",
            "must",
            "pays",
            "regenerates",
            "returns",
            "sacrifices",
            "taps",
            "untaps",
        }
        lead_words = {"when", "whenever", "if", "as", "or", "then"}
        return bool(
            (following and following[0] in action_words)
            or (len(following) >= 2 and following[1] in action_words)
            or (previous and previous[-1] in lead_words and following and following[0] in action_words)
        )

    @staticmethod
    def _rules_text_support_score(card: dict, raw_text: str) -> int:
        ignored = COMMON_EN_OCR_WORDS | COMMON_PT_OCR_WORDS | {
            "about",
            "also",
            "another",
            "could",
            "these",
            "those",
            "unless",
            "would",
        }
        raw_words = [word for word in normalize_text(raw_text).split() if len(word) >= 4]
        rules_text = " ".join(
            [
                card_faces_text(card, "type_line"),
                card_faces_text(card, "oracle_text"),
                card_faces_text(card, "flavor_text"),
                card.get("artist", ""),
            ]
        )
        support_words = [
            word
            for word in normalize_text(rules_text).split()
            if len(word) >= 5 and word not in ignored
        ]
        support_words = list(dict.fromkeys(support_words))
        return fuzzy_word_count(support_words, raw_words, threshold=88)

    def _best_oracle_support(self, card: dict, raw_text: str) -> int:
        oracle_id = card.get("oracle_id")
        same_oracle = [
            item
            for item in self.cards
            if item.get("lang") == "en" and item.get("oracle_id") == oracle_id
        ]
        candidates = same_oracle or [card]
        return max((self._rules_text_support_score(item, raw_text) for item in candidates), default=0)

    def _exact_english_self_reference_in_text(self, raw_text: str) -> tuple[dict | None, str]:
        if not raw_text:
            return None, ""
        choices = self._en_norm_choices()
        ignored_names = COMMON_EN_OCR_WORDS | COMMON_PT_OCR_WORDS | MTG_KEYWORD_NAME_STOPWORDS
        normalized_raw = normalize_text(raw_text)
        best_card = None
        best_line = ""
        best_score = 0
        lines_seen = 0
        for line in raw_text.splitlines():
            cleaned = clean_ocr_line(line)
            words = normalize_text(cleaned).split()
            if not words or len(words) > 12:
                continue
            lines_seen += 1
            if lines_seen > 90:
                break
            for size in range(min(6, len(words)), 0, -1):
                for index in range(0, len(words) - size + 1):
                    candidate = " ".join(words[index : index + size])
                    if candidate not in choices:
                        continue
                    original_name, card_index = choices[candidate]
                    name_words = normalize_text(original_name).split()
                    if len(name_words) != size:
                        continue
                    distinctive_words = [
                        word
                        for word in name_words
                        if len(word) >= 4 and word not in COMMON_EN_OCR_WORDS and word not in COMMON_PT_OCR_WORDS
                    ]
                    if not distinctive_words:
                        continue
                    if size == 1 and (len(candidate) < 5 or candidate in ignored_names):
                        continue
                    if not self._self_reference_line_context(words, index, size):
                        continue
                    card = self.cards[card_index]
                    support = self._rules_text_support_score(card, raw_text)
                    # Old printings whose oracle text was errata'd ("Justice"
                    # -> "this enchantment") barely overlap the printed text,
                    # so oracle support stays low. A name repeated across the
                    # scan is self-reference evidence on its own.
                    occurrences = len(re.findall(rf"\b{re.escape(candidate)}\b", normalized_raw))
                    min_support = 2 if (size == 1 and occurrences >= 2) else 4
                    if support < min_support:
                        continue
                    score = support * 100 + min(sum(len(word) for word in name_words), 30)
                    if score > best_score:
                        best_score = score
                        best_card = card
                        best_line = cleaned
        if best_card:
            return best_card, best_line
        return None, ""

    def _fuzzy_ocr_text(self, raw_text: str) -> tuple[dict | None, str]:
        if not raw_text or not process or not fuzz:
            return None, ""
        ignored = {
            "colar",
            "imagem",
            "abrir",
            "extrair",
            "foil",
            "copiar",
            "google",
            "docs",
            "cabecalho",
            "salvar",
            "excel",
            "atualizar",
            "base",
            "nome",
            "tipo",
            "raridade",
            "descricao",
            "magia",
            "custo",
            "mana",
            "cor",
            "preco",
            "minimo",
            "colecao",
            "ano",
            "poder",
            "qtd",
            "ocr",
            "erro",
            "changeling",
            "this",
            "card",
            "every",
            "creature",
            "type",
            "is",
        }
        choices = {name: index for name, index in self.name_choices}
        normalized_choices = self._en_norm_choices()
        best_card = None
        best_line = ""
        best_score = 0
        lines_seen = 0
        for line in raw_text.splitlines():
            cleaned = clean_ocr_line(line)
            normalized = normalize_text(cleaned)
            words = [word for word in normalized.split() if word not in ignored and len(word) > 1]
            if len(words) > 8:
                continue
            lines_seen += 1
            if lines_seen > 60:
                break
            if words:
                first_word = words[0]
                if len(words) <= 3 and len(first_word) >= 5 and first_word in normalized_choices:
                    original_name, card_index = normalized_choices[first_word]
                    if len(normalize_text(original_name).split()) == 1 and self._title_extra_words_are_noise(words, 1):
                        best_card = self.cards[card_index]
                        best_line = cleaned
                        best_score = max(best_score, 95)
            if len(words) < 2:
                continue
            for size in range(min(5, len(words)), 1, -1):
                for start in range(0, len(words) - size + 1):
                    candidate = " ".join(words[start : start + size])
                    if len(candidate) < 6 or not any(len(word) >= 4 for word in candidate.split()):
                        continue
                    match = process.extractOne(candidate, choices.keys(), scorer=fuzz.WRatio)
                    if match and match[1] > best_score:
                        best_score = match[1]
                        best_card = self.cards[choices[match[0]]]
                        best_line = cleaned
        if best_card and best_score >= 88:
            return best_card, best_line
        return None, ""

    @staticmethod
    def _load_mtgjson_portuguese() -> dict[str, dict]:
        with gzip.open(MTGJSON_ATOMIC_PATH, "rt", encoding="utf-8") as handle:
            payload = json.load(handle)
        translations = {}
        for english_name, variants in payload.get("data", {}).items():
            if isinstance(variants, dict):
                variants = [variants]
            for variant in variants or []:
                foreign_items = variant.get("foreignData") or []
                portuguese = next(
                    (
                        item
                        for item in foreign_items
                        if item.get("language") in {"Portuguese", "Português", "Portuguese (Brazil)"}
                    ),
                    None,
                )
                if portuguese:
                    translations[normalize_text(english_name)] = {
                        "printed_name": portuguese.get("name", ""),
                        "printed_type_line": portuguese.get("type", ""),
                        "printed_text": portuguese.get("text", ""),
                    }
                    break
        return translations

    def portuguese_for(self, card: dict) -> dict | None:
        oracle_id = card.get("oracle_id") or card.get("name")
        manual = self.mtgjson_pt_by_name.get(normalize_text(card.get("name", "")))
        candidates = self.pt_by_oracle.get(oracle_id, [])
        if not candidates:
            return manual
        same_set = [item for item in candidates if item.get("set") == card.get("set")]
        selected = same_set[0] if same_set else candidates[0]
        if manual and (
            not ((selected or {}).get("printed_type_line") or (selected or {}).get("type_line"))
            or not ((selected or {}).get("printed_text") or card_faces_text(selected or {}, "printed_text"))
        ):
            return manual
        return selected

    @staticmethod
    def _single_translated_face_index(card: dict, pt_name: str, pt_type: str, pt_text: str) -> int | None:
        if card.get("layout") != "transform":
            return None
        faces = card.get("card_faces") or []
        if not faces:
            return None
        name_parts = split_face_values(pt_name)
        if len(name_parts) != len(faces):
            return None
        if len(split_face_values(pt_type)) > 1 or len(split_face_values(pt_text)) > 1:
            return None
        if not (pt_type or pt_text):
            return None
        return 0

    @staticmethod
    def _face_power(face: dict) -> str:
        if face.get("power") and face.get("toughness"):
            return f"{face['power']}/{face['toughness']}"
        return ""

    def to_row(self, card: dict, foil: bool) -> list[str]:
        pt_card = self.portuguese_for(card)
        rarity = card.get("rarity", "")
        rarity_value = bilingual(RARITY_EN.get(rarity, rarity.title()), RARITY_PT.get(rarity, ""))

        pt_name = (pt_card or {}).get("printed_name") or (pt_card or {}).get("name") or ""
        pt_type = (pt_card or {}).get("printed_type_line") or (pt_card or {}).get("type_line") or ""
        pt_text = (pt_card or {}).get("printed_text") or card_faces_text(pt_card or {}, "printed_text")
        en_name = card.get("name", "")
        en_type = card_faces_text(card, "type_line")
        en_text = card_faces_text(card, "oracle_text")
        mana_cost = card_mana_cost(card)
        power = card_power(card)

        face_index = self._single_translated_face_index(card, pt_name, pt_type, pt_text)
        if face_index is not None:
            faces = card.get("card_faces") or []
            face = faces[face_index]
            pt_name_parts = split_face_values(pt_name)
            en_name = face.get("name", en_name)
            en_type = face.get("type_line", en_type)
            en_text = face.get("oracle_text", en_text)
            mana_cost = face.get("mana_cost") or mana_cost
            power = self._face_power(face)
            if face_index < len(pt_name_parts):
                pt_name = pt_name_parts[face_index]

        set_code = (card.get("set") or "").upper()
        set_info = self.sets.get(set_code, {})
        set_name = card.get("set_name") or set_info.get("name") or ""
        released_at = card.get("released_at") or set_info.get("released_at") or ""
        finishes = card.get("finishes") or []
        foil_value = "Sim" if foil or finishes == ["foil"] else "Não"

        row = [
            bilingual(en_name, pt_name),
            bilingual(en_type, pt_type),
            rarity_value,
            bilingual(en_text, pt_text),
            mana_cost,
            card_colors_pt(card),
            "",
            f"{set_name} ({set_code})" if set_name else set_code,
            foil_value,
            released_at[:4] if released_at else "",
            power,
            "1",
        ]
        return row


class MagicExtractorApp:
    def __init__(self, root: Tk) -> None:
        self.root = root
        self.main_thread_id = threading.get_ident()
        self.root.title(f"Magic Extractor v{APP_VERSION}")
        self.root.geometry("1180x760")
        self.db = ScryfallDatabase()
        self.ocr = LocalOcr()
        self.current_image: Image.Image | None = None
        self.preview_ref = None
        self.busy_widgets = []
        self.extraction_id = 0
        self.last_appended_row: int | None = None
        self.field_vars = {field: StringVar() for field in FIELDS}
        self.foil_var = BooleanVar(value=False)
        self.status_var = StringVar(value=f"Pronto. v{APP_VERSION}")
        self.raw_ocr_var = StringVar(value="")
        self._build_ui()
        self._log_ocr_status()

    def _build_ui(self) -> None:
        self.root.columnconfigure(1, weight=1)
        self.root.rowconfigure(0, weight=1)

        left = ttk.Frame(self.root, padding=10)
        left.grid(row=0, column=0, sticky="ns")
        left.rowconfigure(2, weight=1)

        buttons = ttk.Frame(left)
        buttons.grid(row=0, column=0, sticky="ew")

        def add_button(row: int, text: str, command, pady=2):
            button = ttk.Button(buttons, text=text, command=command)
            button.grid(row=row, column=0, sticky="ew", pady=pady)
            self.busy_widgets.append(button)
            return button

        add_button(0, "Colar imagem", self.paste_image)
        add_button(1, "Abrir imagem", self.open_image)
        self.extract_button = add_button(2, "Extrair", self.extract)
        foil_check = ttk.Checkbutton(buttons, text="Foil", variable=self.foil_var)
        foil_check.grid(row=3, column=0, sticky="w", pady=4)
        self.busy_widgets.append(foil_check)
        add_button(4, "Copiar linha", self.copy_row)
        add_button(5, "Copiar Google Docs", self.copy_google_docs)
        add_button(6, "Copiar com cabeçalho", self.copy_with_header)
        add_button(7, "Salvar Excel", self.save_excel)
        add_button(8, "Atualizar base", self.download_database, pady=12)

        self.progress = ttk.Progressbar(left, mode="indeterminate")
        self.progress.grid(row=1, column=0, sticky="ew", pady=(0, 8))
        self.progress.grid_remove()

        self.preview = ttk.Label(left, text="Cole ou abra uma imagem", anchor="center")
        self.preview.grid(row=2, column=0, sticky="nsew", pady=10)

        right = ttk.Frame(self.root, padding=10)
        right.grid(row=0, column=1, sticky="nsew")
        right.columnconfigure(1, weight=1)

        self.result_banner = Label(
            right,
            text="",
            font=("Segoe UI", 11, "bold"),
            anchor="w",
            padx=10,
            pady=8,
            wraplength=820,
            justify="left",
        )
        self.result_banner.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 10))

        for row_index, field in enumerate(FIELDS):
            ttk.Label(right, text=field).grid(row=row_index + 1, column=0, sticky="nw", padx=(0, 8), pady=3)
            if field == "Qtd":
                quantity_frame = ttk.Frame(right)
                quantity_frame.columnconfigure(1, weight=1)
                minus_button = ttk.Button(quantity_frame, text="-", width=3, command=lambda: self.adjust_quantity(-1))
                entry = ttk.Entry(quantity_frame, textvariable=self.field_vars[field])
                plus_button = ttk.Button(quantity_frame, text="+", width=3, command=lambda: self.adjust_quantity(1))
                minus_button.grid(row=0, column=0, sticky="ew", padx=(0, 4))
                entry.grid(row=0, column=1, sticky="ew")
                plus_button.grid(row=0, column=2, sticky="ew", padx=(4, 0))
                quantity_frame.grid(row=row_index + 1, column=1, sticky="ew", pady=3)
                self.busy_widgets.extend([minus_button, plus_button])
                continue
            entry = ttk.Entry(right, textvariable=self.field_vars[field])
            if field == "Descrição magia":
                entry = ttk.Entry(right, textvariable=self.field_vars[field])
            entry.grid(row=row_index + 1, column=1, sticky="ew", pady=3)

        ttk.Label(right, text="OCR bruto").grid(row=len(FIELDS) + 1, column=0, sticky="nw", padx=(0, 8), pady=3)
        self.raw_text = ttk.Label(right, textvariable=self.raw_ocr_var, wraplength=850, justify="left")
        self.raw_text.grid(row=len(FIELDS) + 1, column=1, sticky="ew", pady=3)

        status = ttk.Label(self.root, textvariable=self.status_var, anchor="w", padding=(10, 4))
        status.grid(row=1, column=0, columnspan=2, sticky="ew")

        self.duplicate_alert = Label(
            self.root,
            text="",
            fg="red",
            font=("Segoe UI", 10, "bold"),
            anchor="w",
            padx=10,
        )
        self.duplicate_alert.grid(row=2, column=0, columnspan=2, sticky="ew")

    def _log_ocr_status(self) -> None:
        if self.ocr.available():
            lang_note = "por+eng" if self.ocr.por_available else "eng (instale por no Tesseract para cartas PT)"
            self.log(f"OCR pronto: {self.ocr.tesseract_path} | {lang_note} | v{APP_VERSION}")
        else:
            self.log(f"OCR indisponível: instale o Tesseract para extrair texto da imagem. | v{APP_VERSION}")

    def log(self, message: str) -> None:
        if threading.get_ident() != self.main_thread_id:
            self.root.after(0, lambda: self.log(message))
            return
        self.status_var.set(message)
        self.root.update_idletasks()

    def set_busy(self, busy: bool, message: str = "") -> None:
        if busy:
            if message:
                self.status_var.set(message)
            self.progress.grid()
            self.progress.start(12)
            self.root.configure(cursor="watch")
            state = "disabled"
        else:
            self.progress.stop()
            self.progress.grid_remove()
            self.root.configure(cursor="")
            state = "normal"
        for widget in self.busy_widgets:
            widget.configure(state=state)
        self.root.update_idletasks()

    def clear_results(self) -> None:
        for field in FIELDS:
            self.field_vars[field].set("")
        self.raw_ocr_var.set("")
        self.last_appended_row = None
        self.clear_duplicate_alert()
        self.clear_result_banner()

    def show_result_banner(self, message: str, kind: str = "info") -> None:
        styles = {
            "success": ("#1b5e20", "#e8f5e9"),
            "duplicate": ("#b71c1c", "#ffebee"),
            "info": ("#1565c0", "#e3f2fd"),
        }
        fg, bg = styles.get(kind, styles["info"])
        self.result_banner.configure(text=message, fg=fg, bg=bg)
        self.root.update_idletasks()

    def clear_result_banner(self) -> None:
        self.result_banner.configure(text="", bg="SystemButtonFace", fg="SystemWindowText")

    def show_duplicate_alert(self, row_number: int) -> None:
        message = f"Carta já existe na linha {row_number} de {EXCEL_PATH.name} — não foi adicionada de novo."
        self.duplicate_alert.configure(text=message)
        self.show_result_banner(message, "duplicate")

    def clear_duplicate_alert(self) -> None:
        self.duplicate_alert.configure(text="")

    def check_duplicate_alert(self, name: str) -> None:
        if not name.strip():
            self.clear_duplicate_alert()
            return
        existing_row = find_name_row_in_excel(name)
        if existing_row is not None:
            self.show_duplicate_alert(existing_row)
        else:
            self.clear_duplicate_alert()

    def save_or_alert_duplicate(self, row: list[str], allow_update: bool = False) -> None:
        name = row[0]
        if not name.strip():
            return
        existing_row = find_name_row_in_excel(name)
        if existing_row is not None:
            if allow_update and self.last_appended_row == existing_row:
                update_row_in_excel(existing_row, row)
                self.clear_duplicate_alert()
                self.log(f"Atualizado na linha {existing_row} em {EXCEL_PATH}")
                return
            self.show_duplicate_alert(existing_row)
            return
        self.clear_duplicate_alert()
        append_row_to_excel(row)
        self.last_appended_row = find_name_row_in_excel(name)
        self.log(f"Salvo em {EXCEL_PATH}")
        if self.last_appended_row:
            self.show_result_banner(
                f"Salvo na linha {self.last_appended_row} de {EXCEL_PATH.name}.",
                "success",
            )

    def run_background(self, target, done_message: str | None = None, job_id: int | None = None) -> None:
        def is_current_job() -> bool:
            return job_id is None or job_id == self.extraction_id

        def worker() -> None:
            try:
                target()
                if done_message and is_current_job():
                    self.root.after(0, lambda: self.log(done_message))
            except Exception as exc:
                message = str(exc)
                if is_current_job():
                    self.root.after(0, self.clear_results)
                    self.root.after(0, lambda: self.raw_ocr_var.set(f"Erro: {message}"))
                    self.root.after(0, lambda: messagebox.showerror("Erro", message))
                    self.root.after(0, lambda: self.log(f"Erro: {message}"))
            finally:
                if is_current_job():
                    self.root.after(0, lambda: self.set_busy(False))

        threading.Thread(target=worker, daemon=True).start()

    def paste_image(self) -> None:
        grabbed = ImageGrab.grabclipboard()
        if isinstance(grabbed, Image.Image):
            self.set_image(grabbed)
            return
        if isinstance(grabbed, list) and grabbed:
            self.set_image(Image.open(grabbed[0]))
            return
        messagebox.showwarning("Sem imagem", "Não encontrei imagem na área de transferência.")

    def open_image(self) -> None:
        file_path = filedialog.askopenfilename(
            title="Abrir imagem",
            filetypes=[("Imagens", "*.png;*.jpg;*.jpeg;*.webp;*.bmp"), ("Todos", "*.*")],
        )
        if file_path:
            self.set_image(Image.open(file_path))

    def set_image(self, image: Image.Image) -> None:
        self.current_image = crop_possible_app_screenshot(image.convert("RGB"))
        self.clear_results()
        preview = self.current_image.copy()
        preview.thumbnail((340, 520))
        self.preview_ref = ImageTk.PhotoImage(preview)
        self.preview.configure(image=self.preview_ref, text="")
        self.log("Imagem carregada.")

    def refresh_image_from_clipboard(self) -> bool:
        grabbed = ImageGrab.grabclipboard()
        image = None
        if isinstance(grabbed, Image.Image):
            image = grabbed
        elif isinstance(grabbed, list) and grabbed:
            try:
                image = Image.open(grabbed[0])
            except Exception:
                image = None
        if image is None:
            return False
        self.current_image = crop_possible_app_screenshot(image.convert("RGB"))
        preview = self.current_image.copy()
        preview.thumbnail((340, 520))
        self.preview_ref = ImageTk.PhotoImage(preview)
        self.preview.configure(image=self.preview_ref, text="")
        return True

    def extract(self) -> None:
        if self.current_image is None:
            messagebox.showwarning("Sem imagem", "Cole ou abra uma imagem primeiro.")
            return
        self.extraction_id += 1
        job_id = self.extraction_id
        image_for_job = self.current_image.copy()
        self.last_appended_row = None
        self.clear_results()
        self.raw_ocr_var.set(f"Extraindo dados... v{APP_VERSION}")
        self.set_busy(True, "Extraindo dados da carta...")

        def worker() -> None:
            ocr_result = None
            try:
                self.log("Carregando base local...")
                self.db.load(log=self.log)
                self.log("Rodando OCR local...")
                ocr_result = self.ocr.extract(image_for_job)
                self.log("Procurando carta na base...")
                card, reason = self.db.find(ocr_result)
                if job_id != self.extraction_id:
                    return
                self.root.after(0, lambda: self._finish_extract(job_id, ocr_result, card, reason))
            except Exception as exc:
                if job_id != self.extraction_id:
                    return
                message = str(exc)
                captured_ocr = ocr_result
                self.root.after(0, lambda: self._fail_extract(job_id, message, captured_ocr))

        threading.Thread(target=worker, daemon=True).start()

    def _fail_extract(self, job_id: int, message: str, ocr_result: OcrResult | None = None) -> None:
        if job_id != self.extraction_id:
            return
        self.clear_results()
        self.raw_ocr_var.set(f"Erro: {message}")
        self.log(f"Erro: {message}")
        if ocr_result is not None:
            self.write_debug({}, "", ocr_result)
        messagebox.showerror("Erro", message)
        self.set_busy(False)

    def _finish_extract(self, job_id: int, ocr_result: OcrResult, card: dict, reason: str) -> None:
        if job_id != self.extraction_id:
            return
        try:
            self.write_debug(card, reason, ocr_result)
            row = self.db.to_row(card, self.foil_var.get())
            self.fill_fields(row, ocr_result, reason, job_id)
            existing_row = find_name_row_in_excel(row[0])
            if existing_row is None:
                self.save_or_alert_duplicate(self.row_values())
                saved_row = find_name_row_in_excel(row[0])
                if saved_row:
                    message = f"Salvo na linha {saved_row} de {EXCEL_PATH.name}."
                    self.show_result_banner(message, "success")
                    self.log(message)
            else:
                self.show_duplicate_alert(existing_row)
                self.log(
                    f"Preenchido: {row[0]} | {reason} | "
                    f"Já existe na linha {existing_row} de {EXCEL_PATH.name} — não foi adicionada de novo"
                )
                messagebox.showwarning(
                    "Carta já existe",
                    f"\"{row[0]}\" já está na linha {existing_row} de {EXCEL_PATH.name}.\n\n"
                    "Os campos foram preenchidos, mas a carta não foi adicionada de novo ao Excel.",
                )
        except Exception as exc:
            self._fail_extract(job_id, str(exc), ocr_result)
            return
        self.set_busy(False)

    def write_debug(self, card: dict, reason: str, ocr_result: OcrResult) -> None:
        lines = [
            f"version: {APP_VERSION}",
            f"timestamp: {time.strftime('%Y-%m-%d %H:%M:%S')}",
            f"matched_name: {card.get('name', '')}",
            f"matched_set: {(card.get('set') or '').upper()}",
            f"matched_collector: {card.get('collector_number', '')}",
            f"reason: {reason}",
            f"name_hint: {ocr_result.name_hint}",
            f"namebar_text: {ocr_result.namebar_text}",
            f"set_code_hint: {ocr_result.set_code}",
            f"collector_hint: {ocr_result.collector_number}",
            "",
            "raw_ocr:",
            ocr_result.raw_text,
        ]
        text = "\n".join(lines)
        DEBUG_PATH.write_text(text, encoding="utf-8")
        # Keep a rolling history so failed extractions can be reviewed later
        # (the single-file debug above is overwritten on every extraction).
        try:
            with DEBUG_LOG_PATH.open("a", encoding="utf-8") as handle:
                handle.write(f"\n{'=' * 70}\n{text}\n")
        except Exception:
            pass

    def fill_fields(self, row: list[str], ocr_result: OcrResult, reason: str, job_id: int) -> None:
        if job_id != self.extraction_id:
            return
        self.clear_results()
        for field, value in zip(FIELDS, row):
            self.field_vars[field].set(value)
        raw = ocr_result.raw_text.strip()
        self.raw_ocr_var.set(raw[:1200])
        self.show_result_banner(f"Carta identificada: {row[0]}", "info")
        self.log(f"Preenchido: {row[0]} | {reason}")

    def row_values(self) -> list[str]:
        return [self.field_vars[field].get() for field in FIELDS]

    def adjust_quantity(self, delta: int) -> None:
        current = self.field_vars["Qtd"].get().strip()
        try:
            quantity = int(current)
        except ValueError:
            quantity = 0
        self.field_vars["Qtd"].set(str(max(1, quantity + delta)))

    def copy_row(self) -> None:
        self.root.clipboard_clear()
        self.root.clipboard_append("\t".join(self.row_values()))
        self.show_result_banner("Linha copiada", "success")
        self.log("Linha copiada em formato Google Sheets.")

    def copy_google_docs(self) -> None:
        value = " | ".join(self.row_values())
        self.root.clipboard_clear()
        self.root.clipboard_append(value)
        self.log("Linha copiada em texto simples para Google Docs.")

    def copy_with_header(self) -> None:
        value = "\t".join(FIELDS) + "\n" + "\t".join(self.row_values())
        self.root.clipboard_clear()
        self.root.clipboard_append(value)
        self.log("Cabeçalho e linha copiados.")

    def save_excel(self) -> None:
        row = self.row_values()
        if not any(row):
            messagebox.showwarning("Sem dados", "Extraia ou preencha uma carta primeiro.")
            return
        self.save_or_alert_duplicate(row, allow_update=True)

    def download_database(self) -> None:
        self.run_background(lambda: self.db.download(self.log), "Base atualizada.")


def main() -> None:
    root = Tk()
    app = MagicExtractorApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
